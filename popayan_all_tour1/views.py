# =============================================================================
# IMPORTS
# =============================================================================

import json
import os
import traceback
from datetime import datetime, timedelta
from io import BytesIO

# Django — core
from django.apps import apps
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.core.paginator import Paginator
from django.db import connection
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.middleware.csrf import rotate_token
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.decorators.cache import cache_control, never_cache
from django.views.decorators.http import require_POST, require_http_methods
from django.views.generic import DetailView, ListView

# Django — password reset
from django.contrib.auth.views import (
    PasswordResetCompleteView,
    PasswordResetConfirmView,
    PasswordResetDoneView,
    PasswordResetView,
)

# Reportlab — generación de PDFs con reportlab
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.shapes import Drawing
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Openpyxl — generación de Excel
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# xhtml2pdf — generación de PDFs desde HTML
from xhtml2pdf import pisa

# Proyecto — modelos
from .models import (
    CategoriaNoticia,
    Establecimiento,
    Noticia,
    Resena,
    Roles,
    TipoEstablecimiento,
    Usuario,
)

# Proyecto — formularios
from .forms import (
    CategoriaNoticiaForm,
    CompletarPerfilGoogleForm,
    EditarPerfilForm,
    EstablecimientoForm,
    NoticiaForm,
    RegistroUsuarioForm,
    TipoEstablecimientoForm,
    UsuarioForm,
)

# Proyecto — otros
from .avatares import AVATARES_PREDETERMINADOS
from .services import convert_currency, get_currencies


# =============================================================================
# HELPERS GLOBALES
# =============================================================================

def _es_admin(user):
    """Retorna True si el usuario tiene rol administrador."""
    return (
        hasattr(user, 'rol') and
        user.rol is not None and
        user.rol.rol.strip().lower() == 'administrador'
    )


def _json_error(msg, status=400):
    """Respuesta JSON de error estándar."""
    return JsonResponse({'success': False, 'error': msg}, status=status)


def _json_ok(data=None, msg='OK'):
    """Respuesta JSON de éxito estándar."""
    payload = {'success': True, 'message': msg}
    if data:
        payload.update(data)
    return JsonResponse(payload)


def get_client_ip(request):
    """Obtiene la IP real del cliente considerando proxies."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0]
    return request.META.get('REMOTE_ADDR')


# =============================================================================
# HELPERS — ESTABLECIMIENTOS
# =============================================================================

def get_establecimientos_por_tipo(nombre_tipo):
    """Retorna queryset de establecimientos activos filtrados por tipo."""
    return Establecimiento.objects.filter(
        tipo__nombre__iexact=nombre_tipo,
        activo=True
    )


def get_tipo_obj(nombre_tipo):
    """Retorna el objeto TipoEstablecimiento por nombre (case-insensitive)."""
    return TipoEstablecimiento.objects.filter(nombre__iexact=nombre_tipo).first()


# =============================================================================
# HELPERS — PDF
# =============================================================================

def agregar_logo_pdf(story, styles):
    """
    Agrega el logo del proyecto al inicio de un PDF.
    Si no existe el archivo, muestra el nombre del proyecto como texto.
    Ruta esperada: static/img/home/logo.png
    """
    try:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'home', 'logo.png')
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=2 * inch, height=2 * inch)
            logo.hAlign = 'CENTER'
            story.append(logo)
        else:
            logo_text = Paragraph(
                '<b>POPAYÁN ALL TOUR</b>',
                ParagraphStyle(
                    'LogoText',
                    parent=styles['Normal'],
                    fontSize=20,
                    textColor=colors.HexColor('#667eea'),
                    alignment=TA_CENTER,
                )
            )
            story.append(logo_text)
        story.append(Spacer(1, 20))
    except Exception as e:
        print(f"Error al cargar logo: {e}")


def link_callback(uri, rel):
    """
    Convierte URIs de HTML en rutas absolutas del sistema de archivos.
    Requerido por xhtml2pdf para resolver imágenes y estilos en PDFs.
    """
    # Quitar prefijos de URL estático/media
    if uri.startswith(settings.STATIC_URL):
        path = uri.replace(settings.STATIC_URL, "")
    elif uri.startswith(settings.MEDIA_URL):
        path = uri.replace(settings.MEDIA_URL, "")
    else:
        path = uri

    # Buscar en STATIC_ROOT
    if settings.STATIC_ROOT and os.path.exists(settings.STATIC_ROOT):
        full_path = os.path.join(settings.STATIC_ROOT, path)
        if os.path.isfile(full_path):
            return full_path

    # Buscar en carpeta static del proyecto
    full_path = os.path.join(settings.BASE_DIR, 'static', path)
    if os.path.isfile(full_path):
        return full_path

    # Buscar en MEDIA_ROOT
    if settings.MEDIA_ROOT:
        full_path = os.path.join(settings.MEDIA_ROOT, path)
        if os.path.isfile(full_path):
            return full_path

    print(f"⚠️ Advertencia: No se encontró el archivo: {uri}")
    return uri


# =============================================================================
# VISTAS PÚBLICAS — Autenticación
# =============================================================================

def registro(request):
    """Registro de nuevos usuarios."""
    if request.method == "POST":
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Registro exitoso! Ahora puedes iniciar sesión.')
            return redirect("login")
    else:
        form = RegistroUsuarioForm()

    return render(request, "registro/registro.html", {
        "form": form,
        "avatares": AVATARES_PREDETERMINADOS,
    })


@never_cache
@cache_control(no_cache=True, must_revalidate=True, no_store=True)
def login_view(request):
    """Login con email y contraseña. Usa PRG para evitar reenvío del form."""
    if request.method == "POST":
        email = request.POST.get("email", "").strip()
        password = request.POST.get("password", "")
        user = authenticate(request, email=email, password=password)
        if user is not None and user.is_active:
            auth_login(request, user)
            rotate_token(request)
            return redirect("redirect_by_role")
        messages.error(request, "Correo o contraseña incorrectos")
        return redirect("login")

    return render(request, "login/login.html")


@require_POST
@login_required
def logout_view(request):
    """Cierra la sesión del usuario autenticado."""
    logout(request)
    messages.success(request, "Sesión cerrada correctamente.")
    return redirect('login')


@login_required
def completar_perfil_google(request):
    """
    Solicita datos adicionales a usuarios que se registraron con Google
    y aún tienen valores de marcador de posición.
    """
    from datetime import date
    user = request.user

    datos_completos = (
        user.fecha_nacimiento != date(2000, 1, 1)
        and user.telefono
        and user.direccion
        and not user.identificacion.startswith('google-')
    )
    if datos_completos:
        return redirect('home')

    if request.method == 'POST':
        form = CompletarPerfilGoogleForm(request.POST)
        if form.is_valid():
            user.fecha_nacimiento = form.cleaned_data['fecha_nacimiento']
            user.telefono = form.cleaned_data['telefono']
            user.direccion = form.cleaned_data['direccion']
            user.identificacion = form.cleaned_data['identificacion']
            user.save()
            return redirect('home')
    else:
        form = CompletarPerfilGoogleForm()

    return render(request, 'completar_perfil_google.html', {'form': form})


# =============================================================================
# CLASES — Password Reset (vistas personalizadas)
# =============================================================================

class CustomPasswordResetView(PasswordResetView):
    template_name = "password/password_reset.html"
    email_template_name = "password/password_reset_email.txt"
    html_email_template_name = "password/password_reset_email.html"
    subject_template_name = "password/password_reset_subject.txt"


class CustomPasswordResetDoneView(PasswordResetDoneView):
    template_name = 'password/password_reset_done.html'


class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = 'password/password_reset_confirm.html'
    success_url = reverse_lazy('password_reset_complete')


class CustomPasswordResetCompleteView(PasswordResetCompleteView):
    template_name = 'password/password_reset_complete.html'


# =============================================================================
# VISTAS PÚBLICAS — Páginas generales
# =============================================================================

def home(request):
    return render(request, 'home.html')


def terminos(request):
    return render(request, 'registro/terminosYcondiciones.html')


def entretenimiento(request):
    return render(request, 'entretenimiento.html')


def semanas(request):
    return render(request, 'semanaSanta/semana.html')


def memory(request):
    return render(request, 'juego_de_memoria/index.html')


def creditos(request):
    return render(request, 'juegaso/creditos.html')


def menu(request):
    return render(request, 'juegaso/menu.html')


def juegaso(request):
    return render(request, 'juegaso/juego.html')


# =============================================================================
# VISTAS — Perfil de usuario
# =============================================================================

@login_required
def perfilUser(request):
    """
    Perfil del usuario autenticado.
    Permite actualizar datos personales, avatar/imagen y eliminar la cuenta.
    """
    if request.method == 'POST':
        action = request.POST.get('action')

        # --- Eliminar cuenta ---
        if action == 'delete':
            try:
                username = request.user.nombre_completo
                request.user.delete()
                messages.success(request, f'La cuenta de {username} ha sido eliminada exitosamente.')
                return redirect('login')
            except Exception as e:
                messages.error(request, f'Error al eliminar la cuenta: {str(e)}')
                return redirect('perfilUser')

        # --- Actualizar perfil ---
        form = EditarPerfilForm(
            request.POST,
            request.FILES,
            instance=request.user,
            user=request.user,
        )
        if form.is_valid():
            try:
                usuario = form.save(commit=False)
                avatar_id = request.POST.get('avatar_predeterminado')

                if avatar_id:
                    # Seleccionó avatar predeterminado: eliminar imagen personalizada si existe
                    usuario.avatar_predeterminado = avatar_id
                    if usuario.imagen_perfil:
                        try:
                            if os.path.isfile(usuario.imagen_perfil.path):
                                os.remove(usuario.imagen_perfil.path)
                        except Exception:
                            pass
                        usuario.imagen_perfil = None
                elif 'imagen_perfil' in request.FILES:
                    # Subió imagen personalizada: limpiar avatar predeterminado
                    usuario.avatar_predeterminado = None

                usuario.save()
                messages.success(request, '✅ Perfil actualizado exitosamente')
                return redirect('perfilUser')
            except Exception as e:
                messages.error(request, f'Error al guardar: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')

    else:
        form = EditarPerfilForm(instance=request.user, user=request.user)

    resenas = (
        request.user.resenas
        .select_related('establecimiento', 'establecimiento__tipo')
        .order_by('-fecha_creacion')
    )
    favoritos = (
        request.user.favoritos
        .select_related(
            'establecimiento', 'establecimiento__tipo',
            'noticia', 'noticia__categoria',
        )
        .order_by('-fecha_guardado')
    )

    return render(request, 'perfiluser.html', {
        'form': form,
        'usuario': request.user,
        'avatares': AVATARES_PREDETERMINADOS,
        'resenas': resenas,
        'favoritos': favoritos,
    })


@login_required
def eliminar_imagen_perfil(request):
    """Elimina la imagen de perfil del usuario y la resetea al icono por defecto."""
    if request.method == 'POST':
        try:
            user = request.user
            if user.imagen_perfil:
                if os.path.isfile(user.imagen_perfil.path):
                    os.remove(user.imagen_perfil.path)
                user.imagen_perfil = None
                user.save()
                messages.success(request, '✅ Imagen de perfil eliminada correctamente')
            else:
                messages.info(request, 'No hay imagen de perfil para eliminar')
        except Exception as e:
            messages.error(request, f'Error al eliminar la imagen: {str(e)}')

    return redirect('perfilUser')


# =============================================================================
# VISTAS — Redirección por rol
# =============================================================================

@login_required
def redirect_by_role(request):
    """
    Redirige al usuario al panel correspondiente según su rol:
    - empresario  → panel de empresario con sus establecimientos
    - turista     → home
    - administrador → dashboard de administrador
    """
    context = {'usuario': request.user}

    def handle_empresario():
        if not hasattr(request.user, 'tipo_establecimiento') or not request.user.tipo_establecimiento:
            context['error'] = 'Empresario sin tipo de establecimiento asignado'
            return render(request, 'home.html', context)

        tipo_obj = request.user.tipo_establecimiento
        establecimientos_activos = Establecimiento.objects.filter(
            empresario=request.user, activo=True, tipo=tipo_obj
        ).prefetch_related(
            Prefetch(
                'resenas',
                queryset=Resena.objects.select_related('usuario').order_by('-fecha_creacion'),
                to_attr='resenas_lista'
            )
        )
        establecimientos_inactivos = Establecimiento.objects.filter(
            empresario=request.user, activo=False, tipo=tipo_obj
        )
        estadisticas = [
            {
                'establecimiento': e,
                'visitas_totales': 0, 'visitas_mes': 0,
                'visitas_semana': 0, 'visitas_registradas': 0, 'visitas_anonimas': 0,
                'resenas': e.resenas_lista,
                'total_resenas': len(e.resenas_lista),
                'promedio_calificacion': (
                    round(sum(r.calificacion for r in e.resenas_lista) / len(e.resenas_lista), 1)
                    if e.resenas_lista else 0
                ),
            }
            for e in establecimientos_activos
        ]
        context.update({
            'establecimientos_activos': establecimientos_activos,
            'establecimientos_inactivos': establecimientos_inactivos,
            'estadisticas': estadisticas,
            'total_visitas': 0,
            'tipo_establecimiento': tipo_obj.nombre.lower(),
            'titulo': f"Mis {tipo_obj.nombre}",
        })
        return render(request, 'vista_Empresario/V_empre.html', context)

    try:
        if not hasattr(request.user, 'rol') or not request.user.rol:
            context['error'] = 'Usuario sin rol asignado'
            return render(request, 'home.html', context)

        user_role = request.user.rol.rol.strip().lower()

        if user_role == 'empresario':
            return handle_empresario()
        elif user_role == 'turista':
            return render(request, 'home.html', context)
        elif user_role == 'administrador':
            return redirect('dashboard_administrador')
        else:
            context['error'] = f'Rol no reconocido: {user_role}'
            return render(request, 'home.html', context)

    except AttributeError as e:
        context['error'] = f'Error de configuración de usuario: {str(e)}'
        return render(request, 'home.html', context)
    except Exception as e:
        traceback.print_exc()
        context['error'] = f'Error del sistema: {str(e)}'
        return render(request, 'home.html', context)


# =============================================================================
# VISTAS — Establecimientos (empresario)
# =============================================================================

@login_required
def agregar_establecimiento(request, tipo):
    """Permite a un empresario registrar un nuevo establecimiento de su tipo."""
    if not (hasattr(request.user, 'rol') and request.user.rol and request.user.rol.rol.lower() == 'empresario'):
        messages.error(request, 'No tienes permisos para agregar establecimientos.')
        return redirect('home')

    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    if not tipo_obj:
        messages.error(request, "Tipo de establecimiento no válido.")
        return redirect('redirect_by_role')

    if request.method == 'POST':
        form = EstablecimientoForm(request.POST, request.FILES, tipo_id=tipo_obj.id)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.empresario = request.user
            obj.tipo = tipo_obj
            obj.save()
            messages.success(request, f'{tipo_obj.nombre} "{obj.nombre}" agregado exitosamente.')
            return redirect('redirect_by_role')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = EstablecimientoForm(tipo_id=tipo_obj.id)

    return render(request, 'sitios_de_interes/agregar_establecimiento.html', {
        'form': form,
        'titulo': f'Agregar nuevo {tipo_obj.nombre}',
    })


@login_required
def editar_establecimiento(request, tipo, id):
    """Permite a un empresario editar uno de sus establecimientos."""
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    obj = get_object_or_404(Establecimiento, id=id, empresario=request.user, tipo=tipo_obj)

    if request.method == 'POST':
        form = EstablecimientoForm(request.POST, request.FILES, instance=obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'{tipo_obj.nombre} "{obj.nombre}" actualizado exitosamente.')
            return redirect('redirect_by_role')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = EstablecimientoForm(instance=obj)

    return render(request, 'sitios_de_interes/agregar_establecimiento.html', {
        'form': form,
        'titulo': f'Editar {obj.nombre}',
        'objeto': obj,
    })


@login_required
def eliminar_establecimiento(request, tipo, id):
    """Soft delete: desactiva un establecimiento (activo=False)."""
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    obj = get_object_or_404(Establecimiento, id=id, empresario=request.user, tipo=tipo_obj)

    if request.method == 'POST':
        obj.activo = False
        obj.save()
        messages.success(request, f'"{obj.nombre}" desactivado exitosamente.')
        return redirect('redirect_by_role')

    return render(request, 'vistasEmpresario/confirmar_eliminar.html', {'objeto': obj, 'tipo': tipo})


@login_required
def reactivar_establecimiento(request, tipo, id):
    """Reactiva un establecimiento previamente desactivado."""
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    obj = get_object_or_404(Establecimiento, id=id, empresario=request.user, activo=False, tipo=tipo_obj)

    if request.method == 'POST':
        obj.activo = True
        obj.save()
        connection.close()
        messages.success(request, f'"{obj.nombre}" reactivado exitosamente.')
        return redirect('redirect_by_role')

    return render(request, 'vistasEmpresario/confirmar_reactivar.html', {'objeto': obj, 'tipo': tipo})


@login_required
def eliminar_permanente_establecimiento(request, tipo, id):
    """Elimina físicamente un establecimiento de la base de datos."""
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    obj = get_object_or_404(Establecimiento, id=id, empresario=request.user, activo=False, tipo=tipo_obj)

    if request.method == 'POST':
        nombre = obj.nombre
        obj.delete()
        connection.close()
        messages.warning(request, f'"{nombre}" eliminado permanentemente.')
        return redirect('redirect_by_role')

    return render(request, 'vistasEmpresario/confirmar_eliminar_permanente.html', {'objeto': obj, 'tipo': tipo})


# =============================================================================
# VISTAS — Establecimientos (público)
# =============================================================================

def vista_establecimientos(request, tipo):
    """Vista pública de establecimientos por tipo (con botón de registro para empresarios)."""
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    if not tipo_obj:
        tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo.rstrip('s')).first()

    if not tipo_obj:
        messages.error(request, "Tipo de establecimiento no válido.")
        return render(request, 'error.html', {'mensaje': 'Tipo no válido'})

    establecimientos = Establecimiento.objects.filter(tipo=tipo_obj, activo=True)

    botones = []
    if (
        request.user.is_authenticated
        and hasattr(request.user, 'rol')
        and request.user.rol
        and request.user.rol.rol.lower() == 'empresario'
        and request.user.tipo_establecimiento == tipo_obj
    ):
        botones.append({
            'nombre': f'Registrar {tipo_obj.nombre}',
            'url': f'/agregar/{tipo_obj.nombre.lower()}/',
        })

    return render(request, f'sitios_de_interes/{tipo}.html', {
        'establecimientos': establecimientos,
        'botones_establecimiento': botones,
        'tipo': tipo_obj.nombre.lower(),
        'titulo': f"{tipo_obj.nombre} en Popayán",
    })


def listar_establecimientos_publicos(request, tipo):
    """
    Lista pública de establecimientos con reseñas, favoritos y permisos de gestión.
    Incluye prefetch de reseñas para evitar N+1.
    """
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    if not tipo_obj:
        tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo.rstrip('s')).first()

    if tipo_obj:
        objetos = (
            Establecimiento.objects
            .filter(tipo=tipo_obj, activo=True)
            .select_related('tipo', 'empresario')
            .prefetch_related(
                Prefetch(
                    'resenas',
                    queryset=Resena.objects.select_related('usuario').order_by('-fecha_creacion'),
                    to_attr='resenas_recientes',
                )
            )
            .annotate(total_resenas=Count('resenas'))
        )
    else:
        objetos = Establecimiento.objects.none()

    puede_gestionar = False
    if request.user.is_authenticated:
        rol_usuario = getattr(getattr(request.user, 'rol', None), 'rol', '')
        puede_gestionar = rol_usuario.lower() == 'empresario' or request.user.is_staff

    tipo_normalizado = tipo.lower()
    mapa_templates = {
        'hotel': 'hoteles', 'hoteles': 'hoteles',
        'restaurante': 'restaurantes', 'restaurantes': 'restaurantes',
        'museo': 'museos', 'museos': 'museos',
        'iglesia': 'iglesias', 'iglesias': 'iglesias',
    }
    nombre_template = mapa_templates.get(tipo_normalizado, tipo_normalizado)
    tipo_singular = tipo_obj.nombre.lower() if tipo_obj else tipo_normalizado.rstrip('s')

    favoritos_ids = set()
    ya_reseno_ids = set()
    if request.user.is_authenticated:
        favoritos_ids = set(
            request.user.favoritos
            .filter(establecimiento__isnull=False)
            .values_list('establecimiento_id', flat=True)
        )
        ya_reseno_ids = set(
            request.user.resenas
            .values_list('establecimiento_id', flat=True)
        )

    return render(request, f'sitios_de_interes/{nombre_template}.html', {
        'establecimientos': objetos,
        'puede_gestionar': puede_gestionar,
        'titulo': f'{tipo_singular.capitalize()}s en Popayán',
        'tipo': tipo_singular,
        'favoritos_ids': favoritos_ids,
        'ya_reseno_ids': ya_reseno_ids,
    })


def registrar_visita(request, tipo, id):
    """Registra una visita y redirige a la URL externa del establecimiento."""
    tipo_obj = TipoEstablecimiento.objects.filter(nombre__iexact=tipo).first()
    establecimiento = get_object_or_404(Establecimiento, id=id, activo=True, tipo=tipo_obj)
    return redirect(establecimiento.url_mas_info)


# Alias de compatibilidad
def hoteles_view(request):
    return vista_establecimientos(request)


def restaurantes_view(request):
    return vista_establecimientos(request, 'restaurante')


# =============================================================================
# VISTAS — Estadísticas (empresario)
# =============================================================================

@login_required
def estadisticas_establecimiento(request):
    """Panel de estadísticas del empresario (visitas en 0 hasta integrar VisitaEstablecimiento)."""
    if not hasattr(request.user, 'rol') or request.user.rol.rol.lower() != 'empresario':
        messages.error(request, 'No tienes permisos para ver esta sección.')
        return redirect('home')

    tipo_obj = request.user.tipo_establecimiento
    establecimientos = Establecimiento.objects.filter(empresario=request.user, tipo=tipo_obj)
    estadisticas = [
        {'establecimiento': e, 'visitas_totales': 0, 'visitas_mes': 0, 'visitas_semana': 0}
        for e in establecimientos
    ]

    return render(request, 'vista_Empresario/estadisticas.html', {
        'estadisticas': estadisticas,
        'tipo_establecimiento': tipo_obj.nombre.lower() if tipo_obj else '',
        'usuario': request.user,
    })


# =============================================================================
# VISTAS — Dashboard Administrador
# =============================================================================

@login_required
def dashboard_administrador(request):
    """
    Dashboard principal del administrador.
    Carga métricas de establecimientos, usuarios, noticias y datos para gráficas.
    Optimizado con select_related / annotate para evitar consultas N+1.
    """
    if not _es_admin(request.user):
        messages.error(request, 'Acceso restringido.')
        return redirect('home')

    try:
        # ── Establecimientos ─────────────────────────────────────────────
        tipos = TipoEstablecimiento.objects.annotate(
            total=Count('establecimientos', filter=Q(establecimientos__activo=True))
        )
        stats_establecimientos = {t.nombre.lower(): t.total for t in tipos}
        total_establecimientos = sum(stats_establecimientos.values())

        top_establecimientos = [
            {'id': e['id'], 'nombre': e['nombre'], 'tipo': e['tipo__nombre'], 'activo': e['activo'], 'visitas': 0}
            for e in (
                Establecimiento.objects
                .select_related('tipo')
                .values('id', 'nombre', 'tipo__nombre', 'activo')
                .order_by('-activo', 'nombre')  # activos primero
            )
        ]

        # ── Usuarios ─────────────────────────────────────────────────────
        usuarios_qs = (
            Usuario.objects
            .select_related('rol', 'tipo_establecimiento')
            .order_by('nombre_completo')
        )
        usuarios_registrados = usuarios_qs.filter(is_active=True).count()
        usuarios_suspendidos = usuarios_qs.filter(is_active=False).count()

        roles_data = {
            (item['rol__rol'] or 'sin rol').lower(): item['total']
            for item in (
                Usuario.objects
                .filter(is_active=True)
                .values('rol__rol')
                .annotate(total=Count('id'))
            )
        }

        # ── Noticias ─────────────────────────────────────────────────────
        noticias_qs = (
            Noticia.objects
            .select_related('categoria', 'autor')
            .order_by('-fecha_creacion')
        )
        noticias_publicadas = noticias_qs.filter(publicada=True).count()
        noticias_borrador = noticias_qs.filter(publicada=False).count()
        noticias_destacadas = noticias_qs.filter(destacada=True).count()

        # ── Datos para gráficas (JSON) ────────────────────────────────────
        grafica_establecimientos = {
            'labels': [k.title() for k in stats_establecimientos],
            'data': list(stats_establecimientos.values()),
        }
        grafica_usuarios = {
            'labels': ['Total', 'Turistas', 'Empresarios', 'Admins'],
            'data': [
                usuarios_registrados,
                roles_data.get('turista', 0),
                roles_data.get('empresario', 0),
                roles_data.get('administrador', 0),
            ],
        }

        context = {
            'usuario': request.user,
            # Métricas
            'total_establecimientos': total_establecimientos,
            'usuarios_registrados': usuarios_registrados,
            'usuarios_suspendidos': usuarios_suspendidos,
            'usuarios_activos': usuarios_registrados,
            'noticias_publicadas': noticias_publicadas,
            'noticias_borrador': noticias_borrador,
            'noticias_destacadas': noticias_destacadas,
            'roles_data': roles_data,
            # Gráficas
            'grafica_establecimientos': json.dumps(grafica_establecimientos, ensure_ascii=False),
            'grafica_usuarios': json.dumps(grafica_usuarios, ensure_ascii=False),
            'grafica_visitas': json.dumps({'labels': [], 'data': []}, ensure_ascii=False),
            # Tablas
            'top_establecimientos': top_establecimientos,
            'todos_usuarios': usuarios_qs,
            'todos_roles': Roles.objects.all(),
            'todos_tipos': tipos,
            'todas_categorias': CategoriaNoticia.objects.all().order_by('nombre'),
            'noticias_recientes': noticias_qs[:20],
            # Formularios inline
            'estab_form': EstablecimientoForm(),
            'noticia_form': NoticiaForm(),
            'tipo_form': TipoEstablecimientoForm(),
            'cat_form': CategoriaNoticiaForm(),
        }
        return render(request, 'ciudadano/dashboard.html', context)

    except Exception as exc:
        traceback.print_exc()
        messages.error(request, f'Error al cargar el dashboard: {exc}')
        return redirect('home')


# =============================================================================
# VISTAS — Gestión de usuarios (administrador)
# =============================================================================

@login_required
def crear_rol(request):
    """Crea un nuevo rol si no existe ya."""
    if not _es_admin(request.user):
        messages.error(request, 'Sin permisos.')
        return redirect('home')

    if request.method == 'POST':
        nombre = request.POST.get('rol', '').strip()
        if nombre:
            if Roles.objects.filter(rol__iexact=nombre).exists():
                messages.warning(request, f'El rol "{nombre}" ya existe.')
            else:
                Roles.objects.create(rol=nombre)
                messages.success(request, f'Rol "{nombre}" creado exitosamente.')
        else:
            messages.error(request, 'El nombre del rol no puede estar vacío.')

    return redirect('dashboard_administrador')


@login_required
def crear_tipo_establecimiento(request):
    """Crea un nuevo tipo de establecimiento."""
    if not _es_admin(request.user):
        messages.error(request, 'Sin permisos.')
        return redirect('home')

    if request.method == 'POST':
        form = TipoEstablecimientoForm(request.POST)
        if form.is_valid():
            tipo = form.save()
            messages.success(request, f'Tipo "{tipo.nombre}" creado exitosamente.')
        else:
            messages.error(request, 'Error al crear el tipo. Revisa los datos.')

    return redirect('dashboard_administrador')


@login_required
def suspender_usuario(request, user_id):
    """Desactiva la cuenta de un usuario."""
    if not _es_admin(request.user):
        messages.error(request, 'Sin permisos.')
        return redirect('home')

    u = get_object_or_404(Usuario, pk=user_id)
    if u == request.user:
        messages.error(request, 'No puedes suspenderte a ti mismo.')
        return redirect('dashboard_administrador')

    u.is_active = False
    u.save()
    messages.warning(request, f'Usuario "{u.nombre_completo}" suspendido.')
    return redirect('dashboard_administrador')


@login_required
def activar_usuario(request, user_id):
    """Reactiva la cuenta de un usuario suspendido."""
    if not _es_admin(request.user):
        messages.error(request, 'Sin permisos.')
        return redirect('home')

    u = get_object_or_404(Usuario, pk=user_id)
    u.is_active = True
    u.save()
    messages.success(request, f'Usuario "{u.nombre_completo}" activado.')
    return redirect('dashboard_administrador')


@login_required
def editar_usuario(request, user_id):
    """Edita rol, tipo de establecimiento y estado activo de un usuario (vía form POST)."""
    if not _es_admin(request.user):
        messages.error(request, 'Sin permisos.')
        return redirect('home')

    u = get_object_or_404(Usuario, pk=user_id)

    if request.method == 'POST':
        rol_id = request.POST.get('rol')
        tipo_id = request.POST.get('tipo_establecimiento') or None
        activo = request.POST.get('is_active') == 'on'

        if rol_id:
            try:
                u.rol = Roles.objects.get(pk=rol_id)
            except Roles.DoesNotExist:
                messages.error(request, 'Rol no encontrado.')
                return redirect('dashboard_administrador')

        u.tipo_establecimiento_id = tipo_id
        u.is_active = activo
        u.save(update_fields=['rol', 'tipo_establecimiento', 'is_active'])
        messages.success(request, f'Usuario "{u.nombre_completo}" actualizado.')

    return redirect('dashboard_administrador')


# =============================================================================
# VISTAS — Noticias (públicas)
# =============================================================================

def noticia(request):
    """
    Lista todas las noticias publicadas con filtro por categoría,
    ordenamiento y paginación de 9 por página.
    """
    noticias = Noticia.objects.filter(
        publicada=True,
        fecha_publicacion__lte=timezone.now(),
    ).select_related('categoria', 'autor')

    categoria_id = request.GET.get('categoria')
    if categoria_id:
        noticias = noticias.filter(categoria_id=categoria_id)

    orden = request.GET.get('orden', '-fecha_publicacion')
    noticias = noticias.order_by(orden)

    paginator = Paginator(noticias, 9)
    page_obj = paginator.get_page(request.GET.get('page'))

    noticias_populares = (
        Noticia.objects
        .filter(publicada=True, fecha_publicacion__lte=timezone.now())
        .order_by('-visitas_totales')[:5]
    )

    return render(request, 'noticia/noticia.html', {
        'noticias': page_obj,
        'page_obj': page_obj,
        'categorias': CategoriaNoticia.objects.filter(activo=True),
        'noticias_populares': noticias_populares,
    })


def noticias_populares_view(request):
    """Muestra las 10 noticias más visitadas."""
    noticias_populares = (
        Noticia.objects
        .filter(publicada=True, fecha_publicacion__lte=timezone.now())
        .order_by('-visitas_totales')[:10]
    )
    return render(request, 'noticia/noticias_populares.html', {
        'noticias': noticias_populares,
        'titulo': 'Noticias Más Populares',
    })


def noticias_por_categoria_view(request, slug):
    """Muestra noticias filtradas por una categoría específica."""
    categoria = get_object_or_404(CategoriaNoticia, slug=slug, activo=True)
    noticias = (
        Noticia.objects
        .filter(categoria=categoria, publicada=True, fecha_publicacion__lte=timezone.now())
        .order_by('-fecha_publicacion')
    )
    return render(request, 'noticia/noticias_categoria.html', {
        'categoria': categoria,
        'noticias': noticias,
    })


# =============================================================================
# CLASES — Noticias (CBV)
# =============================================================================

class ListaNoticiasView(ListView):
    """Vista basada en clase para listar noticias publicadas con filtros y búsqueda."""
    model = Noticia
    template_name = 'noticia/lista_noticias.html'
    context_object_name = 'noticias'
    paginate_by = 9

    def get_queryset(self):
        queryset = (
            Noticia.objects
            .filter(publicada=True, fecha_publicacion__lte=timezone.now())
            .select_related('categoria', 'autor')
        )
        categoria = self.request.GET.get('categoria')
        if categoria:
            queryset = queryset.filter(categoria__slug=categoria)

        busqueda = self.request.GET.get('q')
        if busqueda:
            queryset = queryset.filter(
                Q(titulo__icontains=busqueda) |
                Q(contenido__icontains=busqueda) |
                Q(subtitulo__icontains=busqueda)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categorias'] = CategoriaNoticia.objects.filter(activo=True)
        context['noticias_destacadas'] = Noticia.objects.filter(
            publicada=True,
            destacada=True,
            fecha_publicacion__lte=timezone.now(),
        )[:3]
        return context


class DetalleNoticiaView(DetailView):
    """Vista de detalle de una noticia con relacionadas, populares y estado de favorito."""
    model = Noticia
    template_name = 'noticia/detalle_noticia.html'
    context_object_name = 'object'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'

    def get_queryset(self):
        return (
            Noticia.objects
            .filter(publicada=True)
            .select_related('categoria', 'autor')
            .prefetch_related('imagenes_adicionales')
        )

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        if request.user.is_authenticated:
            self.object.incrementar_visita(request.user)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['noticias_relacionadas'] = (
            Noticia.objects
            .filter(categoria=self.object.categoria, publicada=True)
            .exclude(id=self.object.id)
            .select_related('categoria')[:3]
        )
        context['noticias_populares'] = (
            Noticia.objects
            .filter(publicada=True, fecha_publicacion__lte=timezone.now())
            .order_by('-visitas_totales')
            .select_related('categoria')[:5]
        )
        if self.request.user.is_authenticated:
            context['ya_visitada'] = self.object.ha_visitado(self.request.user)
            context['es_favorito'] = self.request.user.favoritos.filter(
                noticia=self.object
            ).exists()
        return context


# =============================================================================
# VISTAS — Noticias (administración)
# =============================================================================

@login_required
def crear_noticia(request):
    """Crea una nueva noticia. Accesible para administradores y empresarios."""
    if not (
        request.user.is_superuser or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() in ['administrador', 'empresario'])
    ):
        messages.error(request, 'No tienes permisos para crear noticias')
        return redirect('home')

    if request.method == 'POST':
        form = NoticiaForm(request.POST, request.FILES)
        if form.is_valid():
            noticia_obj = form.save(commit=False)
            noticia_obj.autor = request.user
            if not noticia_obj.fecha_publicacion:
                noticia_obj.fecha_publicacion = timezone.now()
            noticia_obj.save()
            messages.success(request, f'¡Noticia "{noticia_obj.titulo}" creada exitosamente!')
            return redirect('detalle_noticia', slug=noticia_obj.slug)
        else:
            messages.error(request, 'Por favor corrige los errores del formulario')
    else:
        form = NoticiaForm()

    return render(request, 'noticia/form_noticia.html', {
        'form': form,
        'titulo_pagina': 'Crear Nueva Noticia',
        'accion': 'Crear',
    })


@login_required
def editar_noticia(request, slug):
    """Edita una noticia existente. Solo el autor o un administrador puede hacerlo."""
    noticia_obj = get_object_or_404(Noticia, slug=slug)

    if not (
        request.user.is_superuser or
        noticia_obj.autor == request.user or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() == 'administrador')
    ):
        messages.error(request, 'No tienes permisos para editar esta noticia')
        return redirect('detalle_noticia', slug=slug)

    if request.method == 'POST':
        form = NoticiaForm(request.POST, request.FILES, instance=noticia_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'Noticia "{noticia_obj.titulo}" actualizada exitosamente!')
            return redirect('detalle_noticia', slug=noticia_obj.slug)
        else:
            messages.error(request, 'Por favor corrige los errores del formulario')
    else:
        form = NoticiaForm(instance=noticia_obj)

    return render(request, 'noticia/form_noticia.html', {
        'form': form,
        'noticia': noticia_obj,
        'titulo_pagina': 'Editar Noticia',
        'accion': 'Actualizar',
    })


@login_required
def eliminar_noticia(request, slug):
    """Elimina una noticia. Solo el autor o un administrador puede hacerlo."""
    noticia_obj = get_object_or_404(Noticia, slug=slug)

    if not (
        request.user.is_superuser or
        noticia_obj.autor == request.user or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() == 'administrador')
    ):
        messages.error(request, 'No tienes permisos para eliminar esta noticia')
        return redirect('detalle_noticia', slug=slug)

    if request.method == 'POST':
        titulo = noticia_obj.titulo
        noticia_obj.delete()
        messages.success(request, f'Noticia "{titulo}" eliminada exitosamente')
        return redirect('mis_noticias')

    return render(request, 'noticia/eliminar_noticia.html', {'noticia': noticia_obj})


@login_required
def mis_noticias(request):
    """
    Panel de noticias.
    Administradores ven todas; empresarios solo las suyas.
    Permite filtrar por categoría y estado.
    """
    if not (
        request.user.is_superuser or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() in ['administrador', 'empresario'])
    ):
        messages.error(request, 'No tienes permisos para acceder a esta página')
        return redirect('home')

    if request.user.is_superuser or (
        hasattr(request.user, 'rol') and request.user.rol.rol.lower() == 'administrador'
    ):
        noticias = Noticia.objects.all().select_related('categoria', 'autor')
    else:
        noticias = Noticia.objects.filter(autor=request.user).select_related('categoria')

    categoria_id = request.GET.get('categoria')
    estado = request.GET.get('estado')

    if categoria_id:
        noticias = noticias.filter(categoria_id=categoria_id)
    if estado == 'publicada':
        noticias = noticias.filter(publicada=True)
    elif estado == 'borrador':
        noticias = noticias.filter(publicada=False)

    noticias = noticias.order_by('-fecha_creacion')
    paginator = Paginator(noticias, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'noticia/mis_noticias.html', {
        'page_obj': page_obj,
        'categorias': CategoriaNoticia.objects.filter(activo=True),
        'total_noticias': noticias.count(),
    })


@login_required
def toggle_publicar_noticia(request, slug):
    """Alterna el estado publicada/borrador de una noticia."""
    noticia_obj = get_object_or_404(Noticia, slug=slug)

    if not (
        request.user.is_superuser or
        noticia_obj.autor == request.user or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() == 'administrador')
    ):
        messages.error(request, 'No tienes permisos para cambiar el estado de esta noticia')
        return redirect('detalle_noticia', slug=slug)

    noticia_obj.publicada = not noticia_obj.publicada
    if noticia_obj.publicada and not noticia_obj.fecha_publicacion:
        noticia_obj.fecha_publicacion = timezone.now()
    noticia_obj.save()

    estado = "publicada" if noticia_obj.publicada else "despublicada"
    messages.success(request, f'Noticia "{noticia_obj.titulo}" {estado} exitosamente')
    return redirect('noticia/mis_noticias')


@login_required
@require_http_methods(['GET', 'POST'])
def crear_categoria(request):
    """Crea una nueva categoría de noticias vía AJAX (JSON) o form POST."""
    if not (
        request.user.is_superuser or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() == 'administrador')
    ):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return _json_error('No tienes permisos para crear categorías', 403)
        messages.error(request, 'No tienes permisos para crear categorías')
        return redirect('home')

    if request.method == 'POST':
        form = CategoriaNoticiaForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            return _json_ok(
                data={
                    'categoria': {
                        'id': categoria.id,
                        'nombre': categoria.nombre,
                        'descripcion': categoria.descripcion or '',
                        'color': categoria.color,
                        'activo': categoria.activo,
                        'noticias_count': 0,
                    }
                },
                msg=f'Categoría "{categoria.nombre}" creada exitosamente.',
            )
        else:
            errores = {field: errors[0] for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errores}, status=400)

    return redirect('dashboard_administrador')

@login_required
def lista_categorias(request):
    """Lista todas las categorías de noticias."""
    categorias = CategoriaNoticia.objects.all().order_by('nombre')
    return render(request, 'noticia/lista_categorias.html', {'categorias': categorias})


# =============================================================================
# VISTAS — Conversor de divisas
# =============================================================================

def converter_view(request):
    """Página del conversor de divisas."""
    currencies = get_currencies()
    return render(request, "converter.html", {"currencies": currencies})


def convert_api(request):
    """API JSON para convertir divisas."""
    if request.method == "POST":
        try:
            amount = float(request.POST.get("amount"))
            from_cur = request.POST.get("from_currency")
            to_cur = request.POST.get("to_currency")
            converted = convert_currency(amount, from_cur, to_cur)
            return JsonResponse({
                "ok": True,
                "converted": converted,
                "to": to_cur,
                "from": from_cur,
                "amount": amount,
            })
        except Exception as e:
            return JsonResponse({"ok": False, "error": str(e)})


# =============================================================================
# VISTAS — Historia de Popayán
# =============================================================================

# --- Datos históricos --------------------------------------------------------

data_por_ano = {
    1537: {
        "ano": 1537,
        "titulo": "Fundación de Popayán",
        "contenido": [
            "Fundada el 13 de enero de 1537 por el conquistador español Sebastián de Belalcázar, su ubicación estratégica entre Quito y Cartagena la convirtió en un punto clave para las rutas comerciales y militares del virreinato. Desde sus inicios, Popayán destacó por su organización administrativa, su influencia eclesiástica y su papel en la expansión de la corona española en América del Sur. Durante el periodo colonial, la ciudad se consolidó como un centro político, religioso y cultural. La llegada de órdenes religiosas como los franciscanos, dominicos y jesuitas permitió la construcción de iglesias, colegios y seminarios, lo que convirtió a Popayán en un bastión del catolicismo y la educación en el Nuevo Reino de Granada. ",
            "Figuras como el propio Belalcázar y otros encomenderos jugaron un rol determinante en el establecimiento del poder colonial, mientras que los pueblos indígenas locales, como los pubenenses, resistieron valientemente antes de ser sometidos a un nuevo orden social. La fundación de Popayán sentó las bases para el desarrollo del suroccidente colombiano y su historia permanece como un testimonio clave del proceso de conquista y colonización en América."
        ],
        "imagenes": {
            "left": "img-historia/1537_left.webp",
            "right": "img-historia/1537_right.webp"
        },
        "personajes": [
            {
                "nombre": "Sebastián de Belalcázar",
                "fecha": "1479–1551",
                "img_fondo": "img-historia/fondo_1.webp",
                "img_sobre": "img-historia/sobre_1.webp",
                "descripcion": "Fue el fundador de Popayán en 1537. Como conquistador español, estableció rutas estratégicas entre Quito y Cartagena, lo que consolidó la presencia española en el suroccidente colombiano."
            },
            {
                "nombre": "Juan de Ampudia",
                "fecha": "1479–1541",
                "img_fondo": "img-historia/fondo_2.webp",
                "img_sobre": "img-historia/sobre_2.webp",
                "descripcion": "Era uno de los capitanes de Belalcázar, participando activamente en la fundación de varias ciudades. Su papel militar fue esencial para controlar la región y someter a las comunidades indígenas."
            },
            {
                "nombre": "Lorenzo de Aldana",
                "fecha": "1508–1571",
                "img_fondo": "img-historia/fondo_3.webp",
                "img_sobre": "img-historia/sobre_3.webp",
                "descripcion": "Ejerciendo el cargo de gobernador interino tras Belalcázar, su administración ayudó a organizar el sistema colonial local y consolidar el poder español en la zona."
            }
        ],
        "datos_curiosos": [
            "El nombre Popayán proviene del cacique indígena Payán, señor del valle donde se asentaron los españoles",
            "La ciudad fue fundada tres veces: primero en Roldanillo, luego en El Tambo, y finalmente en su ubicación actual.",
            "Popayán fue originalmente pensada como un punto intermedio entre Quito y Cartagena, lo que le dio gran valor estratégico",
            "En sus primeros años, el oro de los ríos del Cauca era lavado por indígenas bajo el sistema de encomienda."
        ]
    },
    1601: {
        "ano": 1601,
        "titulo": "Consolidación de la iglesia",
        "contenido": [
            "Esta etapa, marcada por la institucionalización de la Iglesia Católica en la ciudad, estuvo caracterizada por una intensa actividad misionera, educativa y arquitectónica que definió su identidad religiosa y cultural. La fundación de la diócesis de Popayán en 1546 por el papa Paulo III, con la designación de fray Juan del Valle como su primer obispo, marcó el inicio de un proceso de organización eclesiástica que se afianzó en las décadas posteriores. Desde entonces, obispos, frailes y misioneros trabajaron activamente en la evangelización de la población indígena, la edificación de templos y conventos, y la estructuración de un modelo social basado en la moral católica.",
            "Uno de los acontecimientos clave fue la llegada y expansión de distintas órdenes religiosas. Los franciscanos fueron los primeros en establecerse, seguidos por los dominicos, quienes fundaron el convento de Santo Domingo, y más adelante los jesuitas, que construyeron colegios donde se impartía educación en gramática, latín, teología y filosofía. Estos espacios no solo formaban religiosos, sino también criollos e hijos de encomenderos que más adelante ocuparían cargos importantes en la administración colonial. La Iglesia, además, adquirió grandes extensiones de tierra y riquezas a través de donaciones y herencias, lo que le permitió ejercer una influencia política significativa en la región."
        ],
        "imagenes": {
            "right": "img-historia/anio_1601/1601.webp"
        },
        "personajes": [
            {
                "nombre": "Fray Alonso de Zamora",
                "fecha": "1635 -1717 (Aproximación)",
                "img_fondo": "img-historia/anio_1601/fondo_1.png",
                "img_sobre": "img-historia/anio_1601/sobre_1.png",
                "descripcion": "Se data de él como uno de los primeros frailes dominicos que ayudó a establecer el poder de la Iglesia en Popayán, además de promover la evangelización de los indígenas."
            },
            {
                "nombre": "Juan del Valle",
                "fecha": "1500 - 1563",
                "img_fondo": "img-historia/anio_1601/fondo_2.png",
                "img_sobre": "img-historia/anio_1601/sobre_2.png",
                "descripcion": "Fue el primer obispo de Popayán (1546), y aunque anterior a 1601, su legado perduró al estructurar la diócesis y sentar bases para la educación religiosa.",
            },
        ],
        "datos_curiosos": [
            "Fue una de las primeras ciudades con una diócesis propia en el Nuevo Reino de Granada, desde 1546.",
            "Los franciscanos, dominicos y jesuitas compitieron por construir las iglesias más suntuosas, muchas de las cuales aún existen.",
            "En esa época, los misioneros viajaban hasta el Amazonas desde Popayán para evangelizar pueblos indígenas.",
            "Algunas familias criollas donaban grandes fortunas a la Iglesia para asegurar prestigio y poder local.",
        ]
    },
    1701: {
        "ano": 1701,
        "titulo": "Auge económico y minero",
        "contenido": [
            "Popayán vivió un periodo de gran esplendor económico durante el siglo XVIII, consolidándose como uno de los centros más importantes del Virreinato gracias a la minería de oro y al comercio. Su ubicación estratégica en la ruta entre Quito y Cartagena favoreció el tránsito de mercancías, metales preciosos y viajeros, convirtiéndola en un eje clave del suroccidente del virreinato. Las élites criollas, enriquecidas por la minería en Barbacoas y el Chocó, construyeron fastuosas casonas, templos y capillas, que aún hoy conservan el estilo colonial característico de la ciudad.",
            "Este auge económico permitió el desarrollo de una vida cultural y social sofisticada. Las familias aristocráticas promovieron la educación y el arte, y su influencia se hizo sentir en todos los ámbitos de la vida colonial. Aunque profundamente desigual, esta etapa marcó el crecimiento urbano de Popayán, sentando las bases de su arquitectura, su patrimonio y su posición como símbolo de poder y tradición en el suroccidente colombiano."
        ],
        "imagenes": {
            "left": "img-historia/anio_1701/1701_e.webp",
            "right": "img-historia/anio_1701/1701_right.webp"
        },
        "personajes": [
            {
                "nombre": "Antonio de la Torre y Miranda",
                "fecha": "1734  - 1805",
                "img_fondo": "img-historia/anio_1701/fondo_1.webp",
                "img_sobre": "img-historia/anio_1701/sobre_1.webp",
                "descripcion": "Fue un encomendero y empresario criollo destacado que impulsó la minería en la región del Cauca, enriqueciendo a la élite local."
            },
            {
                "nombre": "Francisco Antonio de Arboleda Salazar",
                "fecha": "1732  - 1793",
                "img_fondo": "img-historia/anio_1701/fondo_2.png",
                "img_sobre": "img-historia/anio_1701/sobre_2.png",
                "descripcion": "Fue un hacendado, militar y político neogranadino influyente de una familia poderosa de Popayán. Participó en la política colonial el cuál consolidó el poder de las élites criollas.",
            },
            {
                "nombre": "José Ignacio de Pombo",
                "fecha": "1761  - 1812",
                "img_fondo": "img-historia/anio_1701/fondo_3.png",
                "img_sobre": "img-historia/anio_1701/sobre_3.png",
                "descripcion": "Comerciante y político que pertenecía a una de las familias fundadoras. Su actividad económica fortaleció la ciudad como centro minero.",
            }
        ],
        "datos_curiosos": [
            "La ciudad fue escenario de tensiones entre realistas y patriotas, con figuras como Camilo Torres y José María Obando.",
            "Muchos próceres y líderes de la independencia nacieron o estudiaron en Popayán, como Francisco José de Caldas.",
            "La élite tradicionalmente apoyaba al rey, sin embargo, con estos sucesos apoyó a la causa libertadora.",
            "La ciudad sufrió saqueos y represalias en las guerras de independencia.",
        ]
    },
    1801: {
        "ano": 1801,
        "titulo": "Popayán en la independecia",
        "contenido": [
            "En esta ciudad nacieron figuras históricas de gran trascendencia, como Camilo Torres, sacerdote y líder revolucionario, y Francisco José de Caldas, científico, ingeniero y patriota. Ambos fueron esenciales en la lucha por la libertad y participaron activamente en los eventos que marcaron la independencia del país. El fervor patriótico que caracterizó a los habitantes de Popayán impulsó numerosas acciones para lograr la separación de España. Durante los años de la independencia, Popayán fue escenario de intensos enfrentamientos armados, pues su posición como centro político y militar la convirtió en un objetivo estratégico tanto para patriotas como para realistas.",
            "Uno de los momentos más críticos ocurrió en 1820, cuando Simón Bolívar envió al general José María Obando a liberar el Cauca. La ciudad fue nuevamente disputada en sangrientos combates, y Popayán, dividida entre partidarios del rey y defensores de la república, sufrió saqueos, incendios y profundas fracturas sociales. Las calles coloniales, hoy tranquilas y patrimoniales, fueron entonces testigos de luchas callejeras, arrestos masivos y persecuciones políticas. A pesar de la violencia y las transformaciones que sufrió, Popayán se mantuvo como un importante núcleo de resistencia y pensamiento revolucionario. Los ideales de libertad germinaron con fuerza en sus claustros, colegios y tertulias intelectuales, y su legado sigue siendo un pilar fundamental en la historia de la independencia de Colombia. La sangre derramada en sus plazas y los sacrificios de sus hijos libertarios son parte esencial del espíritu nacional."
        ],
        "imagenes": {
            "left": "img-historia/anio_1801/1801_left.webp",
            "right": "img-historia/anio_1801/1801_right.webp"
        },
        "personajes": [
            {
                "nombre": "Simón Bolivar",
                "fecha": "1783 - 1830",
                "img_fondo": "img-historia/anio_1801/fondo_1.webp",
                "img_sobre": "img-historia/anio_1801/sobre_1.webp",
                "descripcion": "Simón Bolívar fue clave en la independencia de Colombia, liderando batallas como la de Boyacá en 1819. Su lucha y visión por una América Latina unida lo convirtieron en el principal impulsor de la libertad en la región."
            },
            {
                "nombre": "Antonio Nariño",
                "fecha": "1765 - 1823",
                "img_fondo": "img-historia/anio_1801/fondo_2.webp",
                "img_sobre": "img-historia/anio_1801/sobre_2.webp",
                "descripcion": 'Conocido como "El Precursor", tradujo y difundió los derechos del hombre, promoviendo ideas republicanas y de libertad. Su valentía y compromiso lo llevaron a ser uno de los primeros en enfrentar el dominio español en el país.',
            },
            {
                "nombre": "Tomás Cipriano de Mosquera",
                "fecha": "1798 - 1878",
                "img_fondo": "img-historia/anio_1801/fondo_3.webp",
                "img_sobre": "img-historia/anio_1801/sobre_3.webp",
                "descripcion": "Líder de importantes reformas como la abolición de los diezmos y la desamortización de bienes eclesiásticos, promovió la modernización del Estado y la defensa de la soberanía nacional. Su firme carácter y visión lo convirtieron en un actor fundamental en la consolidación de la República.",
            }
        ],
        "datos_curiosos": [
            "La ciudad fue escenario de tensiones entre realistas y patriotas, con figuras como Camilo Torres y José María Obando.",
            "Muchos próceres y líderes de la independencia nacieron o estudiaron en Popayán, como Francisco José de Caldas.",
            "La élite tradicionalmente apoyaba al rey, sin embargo, con estos sucesos apoyó a la causa libertadora.",
            "La ciudad sufrió saqueos y represalias en las guerras de independencia.",
        ]
    },
    1831: {
        "ano": 1831,
        "titulo": "Fin de la Gran Colombia",
        "contenido": [
            "Para Popayán, una ciudad con fuerte tradición política y conservadora, representó un momento de gran agitación ya que había sido centro del poder colonial y que, tras la independencia, se encontró en medio de profundas transformaciones políticas. La disolución de la Gran Colombia, el ambicioso proyecto integracionista de Simón Bolívar, trajo consigo una ruptura en el orden político que afectó directamente la estructura territorial y el rol que Popayán había desempeñado hasta entonces. La ciudad, que había pertenecido al Departamento del Cauca dentro de esa república, pasó a formar parte de la Nueva Granada, en un proceso cargado de tensiones ideológicas y disputas por el poder regional.",
            "En las calles de Popayán, el pueblo vivía con incertidumbre. El final del proyecto de la Gran Colombia no solo implicaba un nuevo mapa político, sino también una reorganización de los impuestos, la justicia, el comercio y las lealtades militares. La ciudad mantenía su arquitectura colonial y su estructura social jerárquica, pero ya se vislumbraban los conflictos que marcarían el siglo XIX: guerras civiles, disputas entre caudillos regionales y la lucha entre Iglesia y Estado."
        ],
        "personajes": [
            {
                "nombre": "José Hilario López",
                "fecha": "1798 - 1869",
                "img_fondo": "img-historia/anio_1831/fondo_1.webp",
                "img_sobre": "img-historia/anio_1831/sobre_1.webp",
                "descripcion": "Nacido en Popayán en 1798, fue presidente de Colombia y líder liberal. Participó en las guerras de independencia desde joven. Como presidente, abolió la esclavitud en 1851. Promovió reformas agrarias y educativas. Representó la transición del poder desde Popayán hacia un Estado más moderno."
            },
            {
                "nombre": "Julio Arboleda Pombo",
                "fecha": "1817 - 1862",
                "img_fondo": "img-historia/anio_1831/fondo_2.webp",
                "img_sobre": "img-historia/anio_1831/sobre_2.webp",
                "descripcion": "Poeta, político y militar conservador nacido en 1817 en Popayán. Defensor del orden tradicional, fue presidente del Estado Soberano del Cauca. También dirigió fuerzas en guerras civiles. Su obra literaria y liderazgo político influyeron en la identidad regional. Murió asesinado en 1862 durante conflictos internos.",
            },
            {
                "nombre": "Manuel María Mosquera y Arboleda",
                "fecha": "1800 - 1882",
                "img_fondo": "img-historia/anio_1831/fondo_3.webp",
                "img_sobre": "img-historia/anio_1831/sobre_3.webp",
                "descripcion": "Fue diplomático, político y arzobispo destacado en el siglo XIX colombiano. Hijo del expresidente Joaquín Mosquera, perteneció a una de las familias más influyentes de la época. Se desempeñó como representante diplomático en varias misiones internacionales y fue designado Arzobispo de Bogotá en 1859.",
            }
        ],
        "datos_curiosos": [
            "Con la disolución de la Gran Colombia, Popayán pasó a ser parte del Estado Soberano del Cauca, uno de los más grandes.",
            "El Estado del Cauca tenía tanto poder que llegó a tener su propia constitución y ejército."
        ]
    },
    1885: {
        "ano": 1885,
        "titulo": "Guerra civil y la centralización del poder",
        "contenido": [
            "La guerra civil de 1885 surgió como reacción a las reformas liberales y al federalismo que habían dominado décadas anteriores. Las élites de Popayán, ligadas fuertemente a la Iglesia y al poder conservador, se resistieron a la pérdida de influencia que trajo consigo el modelo federalista. Durante el conflicto, la ciudad fue escenario de movilizaciones armadas, enfrentamientos y profundas divisiones internas. Muchos de sus ciudadanos se alistaron en las filas conservadoras, defendiendo un modelo centralista que devolviera el control político al gobierno nacional, alineado con la doctrina católica y el orden tradicional.",
            "Tras la victoria del bando conservador, se impuso una nueva constitución en 1886, que eliminó los Estados Soberanos y fortaleció el poder central en Bogotá. Con ello, Popayán perdió parte de su autonomía política, pero conservó su relevancia cultural y religiosa. El clero, las familias influyentes y las instituciones educativas como el Seminario Mayor y los colegios católicos reforzaron su papel en la formación de las nuevas generaciones bajo los valores del orden conservador."
        ],
        "imagenes": {
            "left": "img-historia/anio_1885/1885_a.webp",
            "right": "img-historia/anio_1885/1885_e.webp"
        },
        "personajes": [
            {
                "nombre": "Miguel Arroyo Hurtado",
                "fecha": "1838 - 1890",
                "img_fondo": "img-historia/anio_1885/fondo_1.webp",
                "img_sobre": "img-historia/anio_1885/sobre_1.webp",
                "descripcion": "Participó en la guerra civil de 1885 como líder de fuerzas conservadoras en el Cauca. Tras el conflicto, ocupó cargos regionales en representación del nuevo gobierno central, encarnando el papel que jugaron los militares locales en la consolidación del orden conservador."
            },
            {
                "nombre": "José María Quijano Wallis",
                "fecha": "1870 - 1923",
                "img_fondo": "img-historia/anio_1885/fondo_2.webp",
                "img_sobre": "img-historia/anio_1885/sobre_2.webp",
                "descripcion": 'Representó el pensamiento conservador tradicionalista y fue cercano a las posturas que apoyaban la centralización. Su influencia fue notable en los debates legales y constitucionales que siguieron a la guerra civil.',
            },
            {
                "nombre": "Manuel Antonio Arboleda Scarpetta",
                "fecha": "1847 - 1922",
                "img_fondo": "img-historia/anio_1885/fondo_3.webp",
                "img_sobre": "img-historia/anio_1885/sobre_3.webp",
                "descripcion": "Ejerció como rector de la Universidad del Cauca y participó activamente en la vida intelectual de la ciudad durante las décadas posteriores a la independencia. Durante la guerra civil de 1885, Quijano defendió abiertamente la causa centralista y conservadora, considerando que el federalismo debilitaba la unidad nacional y la moral católica.",
            }
        ],
        "datos_curiosos": [
            "El conflicto provocó el cierre temporal de escuelas y seminarios, pero la Iglesia los retomó rápidamente.",
            "Muchos patojos ricos estudiaban en Europa, pero regresaban para reforzar el modelo colonialista local.",
            "Durante esta época surgieron publicaciones políticas y literarias en Popayán que promovían ideales católicos y orden social.",
        ]
    },
    1937: {
        "ano": 1937,
        "titulo": "Celebración del IV Centenario",
        "contenido": [
            "La celebración del IV Centenario impulsó la recuperación y embellecimiento del centro histórico, reafirmando a Popayán como una de las joyas patrimoniales de Colombia. Se restauraron edificios coloniales, se levantaron monumentos conmemorativos y se promovieron publicaciones académicas que recogieron su historia. Además, este aniversario consolidó el papel de la ciudad como bastión conservador y centro espiritual del suroccidente colombiano, en un momento en que el país atravesaba tensiones sociales y políticas.",
            "Más allá de la festividad, el IV Centenario se convirtió en un símbolo de continuidad entre el pasado y el presente, resaltando la riqueza cultural de Popayán y su vocación intelectual. Fue también una oportunidad para proyectar la ciudad hacia el futuro, celebrando no solo lo que había sido, sino lo que aspiraba a seguir siendo: un referente de tradición, belleza arquitectónica y conciencia histórica."
        ],
        "imagenes": {
            "left": "img-historia/anio_1937/1937_dere.webp",
            "right": "img-historia/anio_1937/1937_dere_2.webp"
        },
        "personajes": [
            {
                "nombre": "Guillermo Valencia",
                "fecha": "1873 - 1943",
                "img_fondo": "img-historia/anio_1937/fondo_1.webp",
                "img_sobre": "img-historia/anio_1937/sobre_1.webp",
                "descripcion": "Su presencia y obra reforzaron el aura intelectual y conservadora de Popayán durante las celebraciones. Era considerado símbolo del refinamiento literario y de la tradición patricia de la ciudad."
            },
            {
                "nombre": "Rafael Maya",
                "fecha": "1897 - 1980",
                "img_fondo": "img-historia/anio_1937/fondo_2.webp",
                "img_sobre": "img-historia/anio_1937/sobre_2.webp",
                "descripcion": 'Participó en la vida cultural de la ciudad en los años 30, y su obra periodística e intelectual se alineaba con el espíritu de exaltación patrimonial e histórica que marcó la conmemoración.',
            },
            {
                "nombre": "Carlos Albán",
                "fecha": "1888 - 1947",
                "img_fondo": "img-historia/anio_1937/fondo_3.webp",
                "img_sobre": "img-historia/anio_1937/sobre_3.webp",
                "descripcion": "Fue parte del movimiento que promovió investigaciones y publicaciones sobre la historia de la ciudad para conmemorar sus 400 años. Su trabajo ayudó a consolidar la memoria histórica que se destacó en las celebraciones.",
            }
        ],
        "datos_curiosos": [
            "Se construyó el puente del Humilladero, símbolo arquitectónico de la ciudad, para conectar la ciudad alta con la baja.",
            "Durante la conmemoración se revivieron costumbres coloniales como los bailes de salón y vestimenta de época.",
            "Guillermo Valencia, además de poeta, fue embajador y candidato presidencial, y su casa hoy es museo histórico.",
            "Popayán era vista como una ciudad de élite, donde pocas familias concentraban poder político y cultural.",
        ]
    },
    1983: {
        "ano": '1983',
        "titulo": "Terremoto del 31 de marzo",
        "contenido": [
            "El terremoto del 31 de marzo de 1983 marcó un antes y un después en la historia de Popayán, dejando una huella profunda tanto en su arquitectura como en la memoria colectiva de sus habitantes. Aquel Jueves Santo, cuando la ciudad se preparaba para una de las celebraciones religiosas más emblemáticas del país, un sismo de magnitud 5.5 sacudió su territorio con una fuerza inesperada. En pocos segundos, gran parte del centro histórico quedó reducido a escombros. Iglesias, casonas coloniales, calles empedradas y edificios patrimoniales, que durante siglos habían resistido el paso del tiempo, se derrumbaron bajo la violencia de la tierra.",
            "El impacto humano fue igualmente devastador: centenares de muertos, miles de heridos y un número significativo de damnificados que perdieron no solo sus hogares, sino también su tranquilidad y seguridad. La ciudad quedó sumida en el caos, pero al mismo tiempo, el desastre despertó una ola de solidaridad nacional e internacional sin precedentes. Arquitectos, historiadores, ingenieros y ciudadanos de todo el país se unieron en un esfuerzo común por reconstruir Popayán, conservando su esencia colonial y su identidad cultural. La tragedia reveló tanto la fragilidad de un patrimonio edificado como la fortaleza de una comunidad decidida a renacer. Gracias a ese espíritu colectivo, Popayán logró recuperar buena parte de su arquitectura tradicional, convirtiéndose en un símbolo de resiliencia urbana y patrimonial."
        ],
        "imagenes": {
            "right": "img-historia/anio_1983/1983.webp"
        },
        "personajes": [
            {
                "nombre": "Gustavo Wilches-Chaux",
                "fecha": "1954 - Actualidad",
                "img_fondo": "img-historia/anio_1983/fondo_1.webp",
                "img_sobre": "img-historia/anio_1983/sobre_1.webp",
                "descripcion": "Fue uno de los primeros en reflexionar profundamente sobre el concepto de ""gestión del riesgo"" a partir de la experiencia del terremoto de 1983. Su pensamiento influyó en políticas de prevención y manejo de desastres no solo en Popayán, sino a nivel nacional."
            },
            {
                "nombre": "Rafael Maya",
                "fecha": "1897 - 1980",
                "img_fondo": "img-historia/anio_1983/fondo_2.webp",
                "img_sobre": "img-historia/anio_1983/sobre_2.webp",
                "descripcion": 'Participó en la vida cultural de la ciudad en los años 30, y su obra periodística e intelectual se alineaba con el espíritu de exaltación patrimonial e histórica que marcó la conmemoración.',
            },
        ],
        "datos_curiosos": [
            "El terremoto ocurrió el Jueves Santo, interrumpiendo por primera vez en siglos las procesiones de Semana Santa.",
            "Arquitectos de todo el país se unieron a la reconstrucción con técnicas coloniales tradicionales.",
            "Se destruyó más del 80% del centro histórico, incluyendo iglesias coloniales y casas patrimoniales.",
            "A raíz del desastre, Popayán desarrolló uno de los planes de restauración patrimonial más ambiciosos de Colombia.",
        ]
    }
}


# --- Vistas de historia por año ----------------------------------------------

def historia(request, ano=1537):
    datos = data_por_ano.get(ano, data_por_ano[1537])
    return render(request, 'historia.html', {'datos': datos})


def historia_1601_view(request, ano=1601):
    datos = data_por_ano.get(ano, data_por_ano[1601])
    return render(request, 'historia_1601.html', {'datos': datos})


def historia_1701_view(request, ano=1701):
    datos = data_por_ano.get(ano, data_por_ano[1701])
    return render(request, 'historia_1701.html', {'datos': datos})


def historia_1801_view(request, ano=1801):
    datos = data_por_ano.get(ano, data_por_ano[1801])
    return render(request, 'historia_1801.html', {'datos': datos})


def historia_1831_view(request, ano=1831):
    datos = data_por_ano.get(ano, data_por_ano[1831])
    return render(request, 'historia_1831.html', {'datos': datos})


def historia_1885_view(request, ano=1885):
    datos = data_por_ano.get(ano, data_por_ano[1885])
    return render(request, 'historia_1885.html', {'datos': datos})


def historia_1937_view(request, ano=1937):
    datos = data_por_ano.get(ano, data_por_ano[1937])
    return render(request, 'historia_1937.html', {'datos': datos})


def historia_1983_view(request, ano=1983):
    datos = data_por_ano.get(ano, data_por_ano[1983])
    return render(request, 'historia_1983.html', {'datos': datos})


# --- PDFs de historia --------------------------------------------------------

def descargar_historia_completa_pdf(request):
    """Genera y descarga un PDF con toda la historia de Popayán."""
    años_disponibles = [1537, 1601, 1701, 1801, 1831, 1885, 1937, 1983]
    datos_completos = [data_por_ano[a] for a in años_disponibles if a in data_por_ano]

    static_base = (
        settings.STATIC_ROOT
        if settings.STATIC_ROOT and os.path.exists(settings.STATIC_ROOT)
        else os.path.join(settings.BASE_DIR, 'static')
    )

    html_string = render_to_string('pdf/historia_completa.html', {
        'datos_completos': datos_completos,
        'titulo': 'Historia Completa de Popayán',
        'STATIC_URL': settings.STATIC_URL,
        'STATIC_ROOT': static_base,
    }, request=request)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Historia_Popayan_Completa.pdf"'

    pisa_status = pisa.CreatePDF(html_string, dest=response, link_callback=link_callback, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=400)
    return response


def descargar_historia_año_pdf(request, ano):
    """Genera y descarga un PDF con la historia de un año específico."""
    if ano not in data_por_ano:
        return HttpResponse("Año no encontrado", status=404)

    static_base = (
        settings.STATIC_ROOT
        if settings.STATIC_ROOT and os.path.exists(settings.STATIC_ROOT)
        else os.path.join(settings.BASE_DIR, 'static')
    )

    html_string = render_to_string('pdf/historia_año.html', {
        'datos': data_por_ano[ano],
        'titulo': f'Historia de Popayán - Año {ano}',
        'STATIC_URL': settings.STATIC_URL,
        'STATIC_ROOT': static_base,
    }, request=request)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Historia_Popayan_{ano}.pdf"'

    pisa_status = pisa.CreatePDF(html_string, dest=response, link_callback=link_callback, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=400)
    return response


# =============================================================================
# VISTAS — Semana Santa / Procesiones
# =============================================================================

def procesiones(request):
    """
    Vista de procesiones de Semana Santa.
    Los pasos comunes se definen una vez y se reutilizan en cada procesión
    usando spread de diccionario ({**paso, 'numero': N}) para cambiar el número.
    """

    # ── Pasos reutilizables (definidos una vez) ───────────────────────────────

    san_juan_evangelista = {
        'numero': 1, 'imagen': 'img/img_se/pasos_24.png',
        'titulo': 'San Juan Evangelista',
        'descripcion': 'El paso de San Juan Evangelista es uno de los más elegantes y armoniosos, se compone de una anda de madera finamente tallada, adornada con detalles dorados, flores blancas y candelabros que iluminan su recorrido nocturno. Su diseño busca reflejar serenidad y pureza, en coherencia con el papel de San Juan como discípulo fiel.',
        'cargueros': '12 hombres', 'peso': '320 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    la_magdalena = {
        'numero': 2, 'imagen': 'img/img_se/pasos_1.png',
        'titulo': 'La Magdalena',
        'descripcion': 'Representa a María Magdalena, seguidora fiel de Jesús. La imagen la muestra en actitud de recogimiento, con expresión de dolor y penitencia. Es una talla colonial de gran belleza, que resalta la devoción femenina en la Pasión.',
        'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
    }
    la_veronica = {
        'numero': 3, 'imagen': 'img/img_se/pasos_2.png',
        'titulo': 'La Verónica',
        'descripcion': 'Paso que representa el momento en que Verónica limpia el rostro de Cristo camino al Calvario. La tradición asegura que la tela conservó la Santa Faz. La imagen transmite ternura y valentía femenina frente al sufrimiento.',
        'cargueros': '12 hombres', 'peso': '350 kg aprox.', 'material': 'Madera tallada y tela natural',
    }
    el_senor_muerto = {
        'numero': 4, 'imagen': 'img/img_se/pasos_3.jpg',
        'titulo': 'El Señor del Huerto',
        'descripcion': 'Evoca a Cristo en oración en el Monte de los Olivos, momento previo a la Pasión. Su expresión refleja entrega y resignación. Es una de las tallas más antiguas de la procesión, símbolo de recogimiento espiritual.',
        'cargueros': '16 hombres', 'peso': '600 kg aprox.', 'material': 'Madera policromada',
    }
    el_prendimiento = {
        'numero': 5, 'imagen': 'img/img_se/pasos_4.jpg',
        'titulo': 'El Prendimiento',
        'descripcion': 'Escena que muestra la captura de Jesús por los soldados romanos en presencia de Judas Iscariote. Destaca por la fuerza dramática de sus figuras y el realismo en los gestos.',
        'cargueros': '20 hombres', 'peso': '700 kg aprox.', 'material': 'Madera policromada',
    }
    la_negacion = {
        'numero': 6, 'imagen': 'img/img_se/pasos_25.jpg',
        'titulo': 'La Negacion',
        'descripcion': 'Representa el momento en que el apóstol Pedro niega conocer a Jesús antes del canto del gallo, cumpliéndose las palabras del Maestro. El paso muestra una composición escénica en la que Pedro, rodeado por soldados y una criada, refleja el miedo y la debilidad humana frente a la fe',
        'cargueros': '12 hombres', 'peso': '340 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    los_azotes = {
        'numero': 7, 'imagen': 'img/img_se/pasos_26.jpg',
        'titulo': 'Los Azotes',
        'descripcion': 'El paso de Los Azotes representa el momento en que Jesús es flagelado por los soldados romanos antes de ser condenado a muerte. Es una de las escenas más impactantes del Martes Santo, ya que simboliza el sufrimiento, la humillación y la fortaleza de Cristo frente al dolor. El anda está elaborada en madera tallada y dorada, con detalles artísticos que resaltan la crudeza de la escena.',
        'cargueros': '12 hombres', 'peso': '350 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    el_senor_caido = {
        'numero': 8, 'imagen': 'img/img_se/pasos_26.jpg',
        'titulo': 'El Señor Caído',
        'descripcion': 'Figura de Cristo desplomado bajo el peso de la cruz, con rostro de sufrimiento y compasión. Es uno de los pasos más venerados del Martes Santo y despierta profunda devoción en los fieles.',
        'cargueros': '20 hombres', 'peso': '650 kg aprox.', 'material': 'Madera policromada',
    }
    el_ecce = {
        'numero': 9, 'imagen': 'img/img_se/pasos_27.jpg',
        'titulo': 'Ecce Homo',
        'descripcion': 'El paso de Ecce Homo representa el instante en que Pilato presenta a Jesús ante el pueblo, después de haber sido azotado y coronado de espinas, diciendo: "Ecce Homo" (del latín: "He aquí el hombre"). Es una de las escenas más solemnes y simbólicas de la procesión del Martes Santo, pues muestra la humillación de Cristo frente a la multitud.',
        'cargueros': '12 hombres', 'peso': '340 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    el_encuentro = {
        'numero': 10, 'imagen': 'img/img_se/pasos_28.jpg',
        'titulo': 'El Encuentro',
        'descripcion': 'El paso de El Encuentro representa el conmovedor momento en que Jesús, cargando la cruz camino al Calvario, se encuentra con su Madre, la Virgen María. Es una de las escenas más humanas y profundas de la procesión del Martes Santo, pues simboliza el dolor compartido entre madre e hijo ante el sufrimiento inevitable.',
        'cargueros': '12 hombres', 'peso': '350 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    el_amo_jesus = {
        'numero': 11, 'imagen': 'img/img_se/pasos_29.png',
        'titulo': 'El Amo Jesus',
        'descripcion': 'El paso de El Amo Jesús representa a Cristo en su camino hacia el Calvario, cargando la cruz con resignación y dignidad. Es una de las imágenes más queridas y veneradas por los payaneses, considerada símbolo de protección, fe y esperanza para la ciudad. Su nombre popular, "El Amo", refleja la devoción del pueblo hacia Jesús Nazareno como Señor y guía espiritual.',
        'cargueros': '12 hombres', 'peso': '360 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    el_senor_perdon = {
        'numero': 12, 'imagen': 'img/img_se/pasos_30.jpg',
        'titulo': 'El Señor del Perdon',
        'descripcion': 'El paso de El Señor del Perdón representa a Jesús mostrando misericordia y compasión hacia la humanidad, incluso en medio de su sufrimiento camino al Calvario. Su expresión serena y su mirada hacia el cielo reflejan la grandeza del perdón divino, un mensaje central dentro de la Semana Santa payanesa. Este paso invita a la reflexión sobre el arrepentimiento, la reconciliación y la paz interior.',
        'cargueros': '12 hombres', 'peso': '340 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    el_crucifijo = {
        'numero': 15, 'imagen': 'img/img_se/pasos_33.png',
        'titulo': 'El Crucifijo',
        'descripcion': 'El paso de El Crucifijo representa el momento central de la Pasión: Jesús ya clavado en la cruz, consumando su sacrificio redentor por la humanidad. Es uno de los pasos más antiguos y respetados del Martes Santo, símbolo de muerte, redención y esperanza eterna. Su presencia impone un silencio profundo a lo largo del recorrido, invitando a la oración y al recogimiento.',
        'cargueros': '12 hombres', 'peso': '350 kg aprox.', 'material': 'Madera tallada y policromada',
    }
    la_dolorosa = {
        'numero': 17, 'imagen': 'img/img_se/pasos_38.jpg',
        'titulo': 'La Dolorosa',
        'descripcion': 'Simboliza el profundo sufrimiento de la Virgen María ante la Pasión y muerte de su Hijo. Es un paso que expresa el dolor maternal, la fortaleza y la fe inquebrantable. Su presencia en la procesión recuerda el acompañamiento silencioso y amoroso de María en los momentos más difíciles de la Pasión.',
        'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
    }
    la_sentencia = {
        'numero': 7, 'imagen': 'img/img_se/pasos_40.png',
        'titulo': 'La sentencia',
        'descripcion': 'Representa el momento en que Jesús es condenado a muerte por Poncio Pilato. Este paso simboliza la injusticia, la presión del poder y la fragilidad del juicio humano frente a la verdad. Es una escena que marca el inicio del camino hacia la crucifixión y destaca la serenidad y dignidad de Cristo ante la condena.',
        'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
    }

    # ── Datos de cada procesión ───────────────────────────────────────────────

    procesiones_data = {

        # ── Domingo de Ramos ──────────────────────────────────────────────────
        'domingo': {
            'titulo': 'Domingo de Ramos',
            'subtitulo': 'Inicio de la Semana Santa',
            'horario': '9:00 AM',
            'num_pasos': '2 Pasos Procesionales',
            'cargueros_totales': 50,
            'descripcion_1': 'El Domingo de Ramos en Popayán inicia la Semana Santa, una tradición desde el siglo XVI que conmemora la entrada triunfal de Jesús a Jerusalén mediante una procesión diurna.',
            'descripcion_2': 'Esta celebración se documenta desde 1556, con procesiones iniciales en jueves y Viernes Santo por encomendadores reales que portaban cruces y realizaban penitencias. Hasta 1900, honraba al Señor del Triunfo sobre un asno, atrayendo feligreses con palmas de pueblos indígenas vecinos como Yanaconas y Puracé; hoy parte del Santuario de Belén hacia la Catedral Basílica.',
            'descripcion_3': 'A las 9:00 o 10:00 a.m., se realiza la "bajada del Amo" con pasos como El Señor Caído, Santo Ecce Homo, San Juan Evangelista y La Magdalena, cargados por fieles en un recorrido de 2 km en forma de cruz latina por el centro histórico. Los participantes llevan ramos de palma y olivo bendecidos, simbolizando victoria y protección espiritual.',
            'pasos': [
                {**el_ecce, 'numero': 1},
                {
                    'numero': 2, 'imagen': 'img/img_se/pasos_58.png',
                    'titulo': 'El señor Caído',
                    'descripcion': 'Figura de Cristo desplomado bajo el peso de la cruz, con rostro de sufrimiento y compasión. Es uno de los pasos más venerados del Martes Santo y despierta profunda devoción en los fieles.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera tallada y policromada',
                },
            ],
        },

        # ── Lunes Santo ───────────────────────────────────────────────────────
        'lunes': {
            'titulo': 'Lunes Santo',
            'subtitulo': 'Misa del Carguero',
            'horario': '8:00 PM - 1:00 AM',
            'num_pasos': '11 Pasos Procesionales',
            'cargueros_totales': 240,
            'descripcion_1': 'El Lunes Santo en Popayán representa principalmente la solemnidad por la institución de la Sagrada Eucaristía, que es el acto central en la Última Cena donde Jesucristo entrega su cuerpo y sangre a sus apóstoles como sacrificio y comunión espiritual.',
            'descripcion_2': 'Esta celebración es una manifestación de fe enfocada en la memoria histórica y espiritual del sacrificio de Cristo, símbolo de unidad y devoción comunitaria, cultivada en Popayán desde el siglo XVI como una tradición religiosa y cultural.',
            'descripcion_3': 'Además, el Lunes Santo simboliza la recuperación de una tradición perdida por más de un siglo, restituida en 2017 por la Junta Procesional del Lunes Santo pese a polémicas, con apoyo de la Alcaldía y la Universidad del Cauca, incorporando nuevos pasos eucarísticos para enriquecer el ciclo UNESCO.',
            'kilometros': 2, 'horas': 4.0, 'calles': 5,
            'pasos': [
                {
                    'numero': 1, 'imagen': 'img/img_se/pasos_59.jpg',
                    'titulo': 'Jesucristo en la Última Cena',
                    'descripcion': 'Escultura central que representa el momento eucarístico donde Jesús instituye la comunión con los apóstoles, simbolizando sacrificio y unidad.',
                    'cargueros': '15 hombres', 'peso': '320 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 2, 'imagen': 'img/img_se/pasos_60.jpg',
                    'titulo': 'Jesucristo en la Última Cena',
                    'descripcion': 'Escultura central que representa el momento eucarístico donde Jesús instituye la comunión con los apóstoles, simbolizando sacrificio y unidad.',
                    'cargueros': '15 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {**la_magdalena, 'numero': 3},
                {**la_veronica, 'numero': 4},
                {**el_senor_muerto, 'numero': 5},
                {**el_prendimiento, 'numero': 6},
                {**la_negacion, 'numero': 7},
                {**la_sentencia, 'numero': 8},
                {**los_azotes, 'numero': 9},
                {**el_senor_caido, 'numero': 10},
                {**el_ecce, 'numero': 11},
            ],
        },

        # ── Martes Santo ──────────────────────────────────────────────────────
        'martes': {
            'titulo': 'Martes Santo',
            'subtitulo': 'Procesión del Señor del Perdón y María Santísima de los Dolores',
            'horario': '8:00 PM - 11:30 PM',
            'num_pasos': '16 Pasos Procesionales',
            'cargueros_totales': 224,
            'descripcion_1': 'El Martes Santo marca el inicio oficial de las grandes procesiones nocturnas de la Semana Santa payanesa. Esta jornada está dedicada a la meditación sobre el perdón divino y el dolor maternal de María. La procesión sale de la Iglesia de San Francisco y recorre las principales calles del centro histórico, creando un ambiente de profunda espiritualidad.',
            'descripcion_2': 'Los cinco pasos que conforman esta procesión narran episodios fundamentales de la Pasión: desde la oración en el huerto hasta el encuentro de Jesús con su Madre Dolorosa. Cada imagen, tallada por maestros de diferentes épocas, representa siglos de devoción y arte religioso colonial.',
            'descripcion_3': 'Esta procesión se caracteriza por su solemnidad y por ser la más íntima de toda la semana, donde la participación ciudadana es masiva pero respetuosa, creando un silencio sagrado que envuelve las calles empedradas de la ciudad blanca.',
            'kilometros': 2.8, 'horas': 3.5, 'calles': 12,
            'mapa_iframe': '<iframe src="https://www.google.com/maps/embed?pb=!1m34!1m12!1m3!1d3986.1917699962496!2d-76.61088362527605!3d2.4430976570875163!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!4m19!3e0!4m5!1s0x8e300310438eaac9%3A0x473fe8cc717e420f!2sIglesia%20de%20San%20Francisco%2C%20Cra.%209%20%234-29%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4433751!2d-76.6085971!4m5!1s0x8e300305549d061f%3A0xbf6f7babea98c6c4!2sCatedral%20Nuestra%20Se%C3%B1ora%20de%20La%20Asunci%C3%B3n%2C%20Cl.%205a%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4412333!2d-76.60656089999999!4m5!1s0x8e300310438eaac9%3A0x473fe8cc717e420f!2sIglesia%20de%20San%20Francisco%2C%20Cra.%209%20%234-29%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4433751!2d-76.6085971!5e0!3m2!1ses-419!2sco!4v1755772269519!5m2!1ses-419!2sco" width="600" height="450" style="border:0;" allowfullscreen="" loading="lazy" referrerpolicy="no-referrer-when-downgrade"></iframe>',
            'pasos': [
                san_juan_evangelista, la_magdalena, la_veronica, el_senor_muerto,
                el_prendimiento, la_negacion, los_azotes, el_senor_caido,
                el_ecce, el_encuentro, el_amo_jesus, el_senor_perdon,
                {
                    'numero': 13, 'imagen': 'img/img_se/pasos_31.jpg',
                    'titulo': 'El Calvario',
                    'descripcion': 'El paso de El Calvario representa el momento culminante del camino de Jesús hacia la crucifixión, cuando llega al monte del Gólgota para ser entregado al sacrificio. Es una de las escenas más solemnes y meditativas de la procesión del Martes Santo, pues simboliza la culminación del sufrimiento humano y la entrega total de Cristo por la redención del mundo.',
                    'cargueros': '12 hombres', 'peso': '360 kg aprox.', 'material': 'Madera tallada y policromada',
                },
                {
                    'numero': 14, 'imagen': 'img/img_se/pasos_32.png',
                    'titulo': 'El Cristo de la Sed',
                    'descripcion': 'El paso de Cristo de la Sed representa el instante en que Jesús, clavado en la cruz, pronuncia una de sus siete palabras: "Tengo sed" (Juan 19:28). Esta escena simboliza no solo la sed física que sufre Cristo en el Calvario, sino también su sed espiritual de amor y salvación por la humanidad. Es uno de los pasos que más invita a la contemplación y al silencio durante la procesión.',
                    'cargueros': '12 hombres', 'peso': '340 kg aprox.', 'material': 'Madera tallada y policromada',
                },
                el_crucifijo,
                {
                    'numero': 16, 'imagen': 'img/img_se/pasos_6.jpeg',
                    'titulo': 'La Virgen de los Dolores',
                    'descripcion': 'Imagen titular de la procesión, muestra a la Virgen María en profundo dolor al pie de la Cruz. Su rostro refleja tristeza y fortaleza materna. Es el paso más importante del Martes Santo, acompañado con música sacra solemne.',
                    'cargueros': '24 hombres', 'peso': '700 kg aprox.', 'material': 'Madera tallada y vestidura bordada en oro',
                },
            ],
            'puntos_interes': [
                {'titulo': 'Salida - Iglesia San Francisco', 'hora': '8:00 PM', 'descripcion': 'Punto de inicio de la procesión. Los pasos salen en orden establecido desde 1566.'},
                {'titulo': 'Calle del Cauca', 'hora': '8:30 PM', 'descripcion': 'Primera calle principal del recorrido, con balcones coloniales decorados.'},
                {'titulo': 'Plaza Mayor - Catedral', 'hora': '9:15 PM', 'descripcion': 'Momento más solemne, bendición frente a la Catedral Basílica.'},
                {'titulo': 'Calle Real', 'hora': '10:00 PM', 'descripcion': 'Tramo con mayor concentración de público y casas coloniales.'},
                {'titulo': 'Regreso - San Francisco', 'hora': '11:30 PM', 'descripcion': 'Finalización de la procesión y recogida de los pasos.'},
            ],
        },

        # ── Miércoles Santo ───────────────────────────────────────────────────
        'miercoles': {
            'titulo': 'Miércoles Santo',
            'subtitulo': 'Procesión del Amo Jesús y la Virgen Dolorosa',
            'horario': '8:00 PM - 12:00 AM',
            'num_pasos': '17 Pasos Procesionales',
            'cargueros_totales': 220,
            'descripcion_1': 'El Miércoles Santo presenta la procesión más emotiva y conmovedora de toda la Semana Santa payanesa. Conocida como la procesión del "Amo Jesús", está dedicada a meditar sobre el amor incondicional de Cristo y el encuentro doloroso con su Madre.',
            'descripcion_2': 'Los seis pasos que conforman esta procesión narran desde el juicio de Pilatos hasta el calvario del Señor, culminando con el encuentro entre Jesús y María en el camino al Gólgota.',
            'descripcion_3': 'Esta procesión se caracteriza por la participación masiva de familias enteras que han mantenido viva la tradición durante generaciones.',
            'kilometros': 3.5, 'horas': 4.0, 'calles': 16,
            'mapa_iframe': '<iframe src="https://www.google.com/maps/embed?pb=!1m36!1m12!1m3!1d996.548991797857!2d-76.60628336246064!3d2.4416832717349912!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!4m21!3e2!4m5!1s0x8e30030562103b71%3A0x685cc1b9b802d58!2sClaustro%20De%20Santo%20Domingo%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4417267!2d-76.6047488!4m3!3m2!1d2.4410146!2d-76.60515199999999!4m3!3m2!1d2.4415608!2d-76.6069327!4m5!1s0x8e30030562103b71%3A0x685cc1b9b802d58!2sClaustro%20De%20Santo%20Domingo%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4417267!2d-76.6047488!5e0!3m2!1ses-419!2sco!4v1755772600241!5m2!1ses-419!2sco" width="600" height="450" style="border:0;" allowfullscreen="" loading="lazy"></iframe>',
            'pasos': [
                san_juan_evangelista, la_magdalena, la_veronica,
                {
                    'numero': 4, 'imagen': 'img/img_se/pasos_34.png',
                    'titulo': 'La Oracion',
                    'descripcion': 'Representa el momento de profunda oración de Jesús en Getsemaní, previo a su captura. Simboliza la fortaleza espiritual, la obediencia y la entrega total a la voluntad de Dios. Es una talla que transmite serenidad y recogimiento, destacando el sentido de sacrificio en la Pasión.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera tallada y policromada',
                },
                el_prendimiento, la_negacion, los_azotes, el_senor_caido, el_ecce, el_encuentro, el_amo_jesus,
                {
                    'numero': 12, 'imagen': 'img/img_se/pasos_35.jpg',
                    'titulo': 'El despojo',
                    'descripcion': 'Representa el momento en que Jesús es despojado de sus vestiduras antes de la crucifixión. Este paso simboliza la humillación y el despojo material, recordando la entrega total de Cristo por la redención del hombre. Es una talla de gran dramatismo y profundidad espiritual dentro del recorrido procesional.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera tallada y policromada',
                },
                {**el_senor_perdon, 'numero': 13},
                {
                    'numero': 14, 'imagen': 'img/img_se/pasos_36.jpg',
                    'titulo': 'El Cristo de la Sed',
                    'descripcion': 'Evoca el instante en que Jesús, ya en la cruz, expresa una de sus Siete Palabras: "Tengo sed". Este paso simboliza el sufrimiento físico de Cristo y, al mismo tiempo, la sed espiritual de justicia y salvación por la humanidad. Su presencia en la procesión invita a la reflexión sobre el sacrificio y la misericordia divina.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera tallada y policromada',
                },
                {
                    'numero': 15, 'imagen': 'img/img_se/pasos_37.jpg',
                    'titulo': 'El cristo de la Agonia',
                    'descripcion': 'Representa a Jesús en los últimos momentos de vida en la cruz, cuando el sufrimiento alcanza su punto más profundo. Este paso simboliza la entrega definitiva, la redención y la culminación del sacrificio por la humanidad. Su mensaje central es la fortaleza espiritual y el amor llevado hasta el extremo.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera tallada y policromada',
                },
                {**el_crucifijo, 'numero': 16},
                la_dolorosa,
            ],
            'puntos_interes': [
                {'titulo': 'Salida - Iglesia Santo Domingo', 'hora': '8:00 PM', 'descripcion': 'Inicio desde el convento dominico, uno de los más antiguos de la ciudad.'},
                {'titulo': 'Plaza de Caldas', 'hora': '8:45 PM', 'descripcion': 'Primera parada oficial, bendición especial del Amo Jesús.'},
                {'titulo': 'Calle Larga - Plaza Mayor', 'hora': '9:30 PM', 'descripcion': 'Momento culminante frente a la Catedral, mayor concentración de fieles.'},
                {'titulo': 'Barrio San Francisco', 'hora': '10:15 PM', 'descripcion': 'Recorrido por las calles más tradicionales del centro histórico.'},
                {'titulo': 'El Callejón del Embudo', 'hora': '11:00 PM', 'descripcion': 'Tramo más estrecho y pintoresco del recorrido.'},
                {'titulo': 'Regreso - Santo Domingo', 'hora': '12:00 AM', 'descripcion': 'Finalización en el punto de origen tras 4 horas de recorrido.'},
            ],
        },

        # ── Jueves Santo ──────────────────────────────────────────────────────
        'jueves': {
            'titulo': 'Jueves Santo',
            'subtitulo': 'Procesión de la Pasión y Muerte de Nuestro Señor',
            'horario': '8:00 PM - 1:00 AM',
            'num_pasos': '17 Pasos Procesionales',
            'cargueros_totales': 240,
            'descripcion_1': 'El Jueves Santo presenta la procesión más solemne y extensa de toda la Semana Santa payanesa.',
            'descripcion_2': 'Esta procesión se caracteriza por su duración de cinco horas y por incluir las imágenes más grandes y pesadas de toda la semana.',
            'descripcion_3': 'El Cristo de la Veracruz, que cierra la procesión, es considerado una de las tallas más perfectas del arte colonial americano.',
            'kilometros': 4.2, 'horas': 5.0, 'calles': 20,
            'mapa_iframe': '<iframe src="https://www.google.com/maps/embed?pb=!1m38!1m12!1m3!1d3986.194751253035!2d-76.60744512527604!3d2.44209310709254!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!4m23!3e2!4m5!1s0x8e300305bc631871%3A0xeff9e7336c0a4255!2sErmita%20de%20Jes%C3%BAs%20Nazareno%2C%20Cl%205%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4401525!2d-76.6028527!4m5!1s0x8e30033b77c3ec87%3A0x33ee2a06c47a4e76!2sPuente%20Del%20Humilladero%2C%20Cra%205%20%231-28%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.444023!2d-76.60508519999999!4m3!3m2!1d2.4415608!2d-76.6069327!4m5!1s0x8e300305bc631871%3A0xeff9e7336c0a4255!2sErmita%20de%20Jes%C3%BAs%20Nazareno%2C%20Cl%205%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4401525!2d-76.6028527!5e0!3m2!1ses-419!2sco!4v1755772716706!5m2!1ses-419!2sco" width="600" height="450" style="border:0;" allowfullscreen="" loading="lazy"></iframe>',
            'pasos': [
                san_juan_evangelista, la_magdalena, la_veronica, el_senor_muerto,
                {
                    'numero': 5, 'imagen': 'img/img_se/pasos_39.png',
                    'titulo': 'El beso de Judas',
                    'descripcion': 'Representa el momento en que Judas Iscariote identifica a Jesús ante los soldados mediante un beso, acto que da inicio a su prendimiento. Este paso simboliza la traición, la fragilidad humana y el contraste entre la lealtad y la corrupción moral. Es una escena clave dentro de la Pasión, recordada por su fuerte carga dramática y espiritual.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {**el_prendimiento, 'numero': 6},
                la_sentencia,
                {**los_azotes, 'numero': 8},
                {
                    'numero': 9, 'imagen': 'img/img_se/pasos_41.png',
                    'titulo': 'La coronacion',
                    'descripcion': 'Representa el momento en que Jesús es coronado con espinas por los soldados romanos, en un acto de burla y humillación. Este paso simboliza el sufrimiento aceptado con humildad, la realeza espiritual de Cristo y la ironía con la que fue tratado antes de la crucifixión. Es una escena que resalta la dignidad y fortaleza de Jesús ante el dolor y la burla.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {**el_ecce, 'numero': 10},
                {
                    'numero': 11, 'imagen': 'img/img_se/pasos_42.png',
                    'titulo': 'La cruz a Cuestas',
                    'descripcion': 'Representa a Jesús avanzando hacia el Calvario mientras carga la cruz. Este paso simboliza el peso del sacrificio, la obediencia y la entrega total por la salvación de la humanidad. También evoca la perseverancia ante el sufrimiento y el camino doloroso que Cristo decide asumir por amor. Es una de las escenas más significativas dentro del recorrido procesional.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                el_senor_perdon,
                {
                    'numero': 13, 'imagen': 'img/img_se/pasos_44.png',
                    'titulo': 'La Crucifixión',
                    'descripcion': 'Representa el momento central de la Pasión: Jesús clavado en la cruz, acompañado de los dos ladrones y custodiado por soldados romanos.',
                    'cargueros': '24 hombres', 'peso': '800 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 14, 'imagen': 'img/img_se/pasos_43.png',
                    'titulo': 'El Calvario',
                    'descripcion': 'Paso que muestra a Cristo crucificado junto a la Virgen María, San Juan Evangelista y María Magdalena al pie de la cruz.',
                    'cargueros': '24 hombres', 'peso': '850 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 15, 'imagen': 'img/img_se/pasos_13.jpg',
                    'titulo': 'El Señor de la Expiración',
                    'descripcion': 'Representa el momento en que Jesús entrega su último aliento en la cruz. Este paso simboliza la culminación del sacrificio redentor, la entrega total y el amor llevado hasta sus últimas consecuencias. Es una escena profundamente espiritual que invita al recogimiento y a la contemplación del acto supremo de entrega de Cristo por la humanidad.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 16, 'imagen': 'img/img_se/pasos_45.png',
                    'titulo': 'El Santo Cristo de la Santa Veracruz',
                    'descripcion': 'Es un paso que presenta a Cristo crucificado bajo la advocación de la Vera Cruz, símbolo de la verdadera cruz donde fue entregada su vida por la humanidad. Representa la solemnidad del sacrificio redentor y la victoria espiritual de Cristo sobre el pecado y la muerte. Es un paso de profunda veneración, tradicional en las procesiones de Popayán por su fuerte carácter devocional e histórico.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                la_dolorosa,
            ],
            'puntos_interes': [
                {'titulo': 'Salida - Iglesia La Ermita', 'hora': '8:00 PM', 'descripcion': 'Inicio desde el templo más antiguo, construido en 1546.'},
                {'titulo': 'Puente del Humilladero', 'hora': '8:30 PM', 'descripcion': 'Cruce simbólico del río Molino, tradición de más de 400 años.'},
                {'titulo': 'Plaza Mayor', 'hora': '10:00 PM', 'descripcion': 'Momento central de la procesión, mayor solemnidad.'},
                {'titulo': 'Calle del Carozo', 'hora': '11:30 PM', 'descripcion': 'Tramo más empinado, mayor esfuerzo de los cargueros.'},
                {'titulo': 'Regreso - La Ermita', 'hora': '1:00 AM', 'descripcion': 'Finalización después de 5 horas de recorrido.'},
            ],
        },

        # ── Viernes Santo ─────────────────────────────────────────────────────
        'viernes': {
            'titulo': 'Viernes Santo',
            'subtitulo': 'Procesión del Santo Entierro',
            'horario': '7:00 PM - 11:00 PM',
            'num_pasos': '13 Pasos Procesionales',
            'cargueros_totales': 156,
            'descripcion_1': 'El Viernes Santo representa el momento más solemne y emotivo de toda la Semana Santa payanesa.',
            'descripcion_2': 'Los siete pasos que conforman esta procesión narran desde la crucifixión hasta el entierro de Jesús.',
            'descripcion_3': 'La procesión del Viernes Santo se distingue por su carácter fúnebre y por la participación de autoridades civiles y militares.',
            'kilometros': 3.0, 'horas': 4.0, 'calles': 14,
            'mapa_iframe': '<iframe src="https://www.google.com/maps/embed?pb=!1m38!1m12!1m3!1d498.27523958011557!2d-76.60693580989974!3d2.439677001586291!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!4m23!3e2!4m5!1s0x8e30031b48914881%3A0xe36da6139c290913!2sIglesia%20San%20Agust%C3%ADn%2C%20Cra.%206%20%23762%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4394036999999997!2d-76.6068201!4m3!3m2!1d2.4415608!2d-76.6069327!4m5!1s0x8e30030ff9207c27%3A0xf76217f9f8464ebe!2sParque%20Caldas%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4418674!2d-76.60627389999999!4m5!1s0x8e30031b48914881%3A0xe36da6139c290913!2sIglesia%20San%20Agust%C3%ADn%2C%20Cra.%206%20%23762%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4394036999999997!2d-76.6068201!5e0!3m2!1ses-419!2sco!4v1755772856950!5m2!1ses-419!2sco" width="600" height="450" style="border:0;" allowfullscreen="" loading="lazy"></iframe>',
            'pasos': [
                {
                    'numero': 1, 'imagen': 'img/img_se/pasos_15.jpg',
                    'titulo': 'La Muerte',
                    'descripcion': 'Representa la figura alegórica de la Muerte, vestida con túnica y portando símbolos fúnebres.',
                    'cargueros': '16 hombres', 'peso': '500 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 2, 'imagen': 'img/img_se/pasos_16.jpg',
                    'titulo': 'María Salomé',
                    'descripcion': 'Figura de una de las mujeres que acompañó a Cristo hasta la crucifixión.',
                    'cargueros': '12 hombres', 'peso': '350 kg aprox.', 'material': 'Madera policromada',
                },
                la_veronica,
                {**la_magdalena, 'numero': 4},
                {
                    'numero': 5, 'imagen': 'img/img_se/pasos_47.png',
                    'titulo': 'El Varón del Martillo',
                    'descripcion': 'Figura alegórica de uno de los sayones encargados de ejecutar la crucifixión.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 6, 'imagen': 'img/img_se/pasos_46.png',
                    'titulo': 'El Varón de las Tenazas',
                    'descripcion': 'Representa a Jesús en la cruz acompañado de los símbolos de la Pasión, entre ellos las tenazas, utilizadas para retirar los clavos al momento de descender su cuerpo. Este paso simboliza la consumación del sacrificio, la humanidad de Cristo y los instrumentos que formaron parte de su martirio. Es una advocación que invita a la reflexión sobre el dolor y la entrega total manifestada en la crucifixión.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 7, 'imagen': 'img/img_se/pasos_19.jpg',
                    'titulo': 'El Santo Cristo',
                    'descripcion': 'Paso central del Viernes Santo, representa a Jesús ya crucificado.',
                    'cargueros': '24 hombres', 'peso': '800 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 8, 'imagen': 'img/img_se/pasos_20.png',
                    'titulo': 'El Descendimiento',
                    'descripcion': 'Representa el momento en que Cristo es bajado de la cruz, asistido por José de Arimatea y Nicodemo.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 9, 'imagen': 'img/img_se/pasos_48.png',
                    'titulo': 'La piedad',
                    'descripcion': 'Representa a la Virgen María sosteniendo en sus brazos el cuerpo de Jesús después de ser descendido de la cruz. Este paso simboliza el dolor profundo de una madre, la compasión y el vínculo inseparable entre María y Cristo en el momento más trágico de la Pasión. Es una escena de recogimiento que inspira reflexión sobre el amor, el sacrificio y la misericordia.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 10, 'imagen': 'img/img_se/pasos_49.png',
                    'titulo': 'El Traslado de Cristo hasta el Sepulcro',
                    'descripcion': 'Representa el momento en que el cuerpo de Jesús, ya descendido de la cruz, es llevado por sus discípulos y seguidores hacia el lugar donde será sepultado. Este paso simboliza respeto, veneración y el último acto de amor hacia Cristo antes de su descanso en el sepulcro. Es una escena que invita al silencio, la reflexión y la contemplación del sacrificio consumado.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {**san_juan_evangelista, 'numero': 11},
                {
                    'numero': 12, 'imagen': 'img/img_se/pasos_50.jpg',
                    'titulo': 'El Santo Sepulcro',
                    'descripcion': 'Representa a Jesús ya yacente dentro del sepulcro, después de haber sido preparado para su entierro. Este paso simboliza el reposo sagrado de Cristo tras culminar su sacrificio, así como la espera silenciosa antes de la Resurrección. Es uno de los pasos más solemnes y venerados, invitando al recogimiento profundo y a la contemplación del misterio de la muerte y la esperanza cristiana.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 13, 'imagen': 'img/img_se/pasos_51.png',
                    'titulo': 'La Virgen de la Soledad',
                    'descripcion': 'Representa a María en el dolor silencioso y profundo que experimenta tras la muerte de su Hijo. Este paso simboliza la soledad, la fidelidad y la fortaleza interior de la Virgen en el momento más desolador de la Pasión. Su figura invita al recogimiento y a la contemplación del amor maternal que permanece firme incluso en medio del sufrimiento.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
            ],
            'puntos_interes': [
                {'titulo': 'Salida - Iglesia San Agustín', 'hora': '7:00 PM', 'descripcion': 'Inicio desde el convento agustino, en absoluto silencio.'},
                {'titulo': 'Calle de la Universidad', 'hora': '7:45 PM', 'descripcion': 'Paso frente a la histórica Universidad del Cauca.'},
                {'titulo': 'Plaza Mayor - Momento Solemne', 'hora': '8:30 PM', 'descripcion': 'Honores militares al Santo Sepulcro, momento más emotivo.'},
                {'titulo': 'Calle Real', 'hora': '9:15 PM', 'descripcion': 'Recorrido por el centro comercial histórico.'},
                {'titulo': 'Regreso - San Agustín', 'hora': '11:00 PM', 'descripcion': 'Finalización con el recogimiento del Santo Sepulcro.'},
            ],
        },

        # ── Sábado Santo ──────────────────────────────────────────────────────
        'sabado': {
            'titulo': 'Sábado Santo',
            'subtitulo': 'Procesión de la Resurrección',
            'horario': '6:00 AM - 8:00 AM',
            'num_pasos': '4 Pasos Procesionales',
            'cargueros_totales': 80,
            'descripcion_1': 'El Sábado Santo cierra la Semana Santa payanesa con la Procesión de la Resurrección, un evento lleno de alegría y esperanza.',
            'descripcion_2': 'La procesión se caracteriza por el repique de campanas, música festiva y la participación masiva de familias con niños.',
            'descripcion_3': 'Esta procesión representa la esperanza cristiana y la vida nueva. A diferencia de las procesiones nocturnas de la semana, se realiza al amanecer.',
            'kilometros': 2.0, 'horas': 2.0, 'calles': 8,
            'mapa_iframe': '<iframe src="https://www.google.com/maps/embed?pb=!1m38!1m12!1m3!1d1185.1029570517421!2d-76.60695166810642!3d2.441903721488239!2m3!1f0!2f0!3f0!3m2!1i1024!2i768!4f13.1!4m23!3e2!4m5!1s0x8e300305549d061f%3A0xbf6f7babea98c6c4!2sCatedral%20Nuestra%20Se%C3%B1ora%20de%20La%20Asunci%C3%B3n%2C%20Cl.%205a%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4412333!2d-76.60656089999999!4m5!1s0x8e30030ff9207c27%3A0xf76217f9f8464ebe!2sParque%20Caldas%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4418674!2d-76.60627389999999!4m3!3m2!1d2.4427116!2d-76.607478!4m5!1s0x8e300305549d061f%3A0xbf6f7babea98c6c4!2sCatedral%20Nuestra%20Se%C3%B1ora%20de%20La%20Asunci%C3%B3n%2C%20Cl.%205a%2C%20Centro%2C%20Popay%C3%A1n%2C%20Cauca!3m2!1d2.4412333!2d-76.60656089999999!5e0!3m2!1ses-419!2sco!4v1755772950630!5m2!1ses-419!2sco" width="600" height="450" style="border:0;" allowfullscreen="" loading="lazy"></iframe>',
            'pasos': [
                {
                    'numero': 1, 'imagen': 'img/img_se/pasos_21.png',
                    'titulo': 'Cirio Pascual',
                    'descripcion': 'Representa la luz de Cristo resucitado, signo de esperanza y vida eterna.',
                    'cargueros': '12 hombres', 'peso': '400 kg aprox.', 'material': 'Madera y cera bendita',
                },
                {
                    'numero': 2, 'imagen': 'img/img_se/pasos_22.png',
                    'titulo': 'María Salomé',
                    'descripcion': 'Una de las santas mujeres que acudió al sepulcro. Figura de fidelidad y devoción.',
                    'cargueros': '18 hombres', 'peso': '600 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 3, 'imagen': 'img/img_se/pasos_23.png',
                    'titulo': 'María la Madre de Santiago',
                    'descripcion': 'Mujer presente en el hallazgo del sepulcro vacío, símbolo de ternura y testimonio.',
                    'cargueros': '18 hombres', 'peso': '600 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 4, 'imagen': 'img/img_se/pasos_52.png',
                    'titulo': 'María Magdalena',
                    'descripcion': 'Representa a María Magdalena como testigo y anunciadora de la Resurrección. Este paso simboliza la esperanza renovada, la alegría espiritual y la misión de proclamar que Cristo ha vencido a la muerte. Su figura encarna la fidelidad y el amor que permanecen más allá del sufrimiento, convirtiéndose en signo de vida nueva y renovación para los creyentes.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 5, 'imagen': 'img/img_se/pasos_53.png',
                    'titulo': 'San Juan',
                    'descripcion': 'Representa a San Juan Evangelista como discípulo amado y testigo de la Resurrección. Simboliza la fidelidad, la esperanza y la alegría del triunfo de Cristo sobre la muerte. Es un paso propio del Sábado Santo, que acompaña al Resucitado en la procesión festiva de ese día.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 6, 'imagen': 'img/img_se/pasos_54.png',
                    'titulo': 'San Pedro',
                    'descripcion': 'Representa al apóstol Pedro como figura de fortaleza y guía de la Iglesia. Simboliza la fe renovada tras la Resurrección, el arrepentimiento y la misión de anunciar el mensaje de Cristo. En el Sábado Santo acompaña al Resucitado como testigo del triunfo definitivo de Jesús.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 7, 'imagen': 'img/img_se/pasos_55.png',
                    'titulo': 'Nuestra Señora de la Pascua',
                    'descripcion': 'Representa a la Virgen María como Madre gozosa ante la Resurrección de su Hijo. Simboliza la alegría plena, la esperanza renovada y el cumplimiento de la promesa de vida nueva. Es un paso característico del Sábado Santo, acompañando al Cristo Resucitado en un ambiente festivo y de celebración.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
                {
                    'numero': 8, 'imagen': 'img/img_se/pasos_56.png',
                    'titulo': 'Nuestro Señor Jesucristo Resucitado',
                    'descripcion': 'Representa a Cristo triunfante después de vencer la muerte, proclamando la vida nueva y la gloria de la Resurrección. Simboliza la victoria, la esperanza y la alegría cristiana que marca el final de la Pasión y el inicio del tiempo pascual. Es el paso central del Sábado Santo y el motivo de la celebración más festiva de la Semana Santa en Popayán.',
                    'cargueros': '12 hombres', 'peso': '300 kg aprox.', 'material': 'Madera policromada',
                },
            ],
            'puntos_interes': [
                {'titulo': 'Salida - Catedral Basílica', 'hora': '6:00 AM', 'descripcion': 'Inicio al amanecer desde el templo mayor de la ciudad.'},
                {'titulo': 'Plaza Mayor - Celebración', 'hora': '6:30 AM', 'descripcion': 'Momento de mayor alegría con repique de campanas.'},
                {'titulo': 'Calle Real', 'hora': '7:00 AM', 'descripcion': 'Recorrido festivo con participación de familias.'},
                {'titulo': 'Regreso - Catedral', 'hora': '8:00 AM', 'descripcion': 'Finalización con Misa de Resurrección.'},
            ],
        },
    }

    return render(request, 'semanaSanta/procesiones.html', {'procesiones': procesiones_data})


# =============================================================================
# VISTAS — Exportación PDF/Excel (administrador)
# =============================================================================

@login_required
def exportar_estadisticas_admin_pdf(request):
    """Genera y descarga un PDF con el resumen del dashboard del administrador."""
    if not _es_admin(request.user):
        messages.error(request, 'No tienes permisos para esta acción.')
        return redirect('home')

    # ── Recopilar datos ───────────────────────────────────────────────────────
    tipos = TipoEstablecimiento.objects.all()
    stats_establecimientos = {
        tipo.nombre.lower(): Establecimiento.objects.filter(tipo=tipo, activo=True).count()
        for tipo in tipos
    }
    total_establecimientos = sum(stats_establecimientos.values())
    usuarios_registrados = Usuario.objects.filter(is_active=True).count()

    roles_dict = {
        (item['rol__rol'] or '').lower(): item['total']
        for item in Usuario.objects.filter(is_active=True).values('rol__rol').annotate(total=Count('id'))
    }

    # Visitas desactivadas (sin modelo VisitaEstablecimiento)
    total_visitas = visitas_mes = visitas_registradas = visitas_invitados = 0

    top_establecimientos = [
        {'nombre': e['nombre'], 'tipo': e['tipo__nombre'], 'visitas': 0}
        for e in (
            Establecimiento.objects.filter(activo=True)
            .values('nombre', 'tipo__nombre')
            .order_by('nombre')[:5]
        )
    ]

    # ── Construir PDF ─────────────────────────────────────────────────────────
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    story = []
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                  textColor=colors.HexColor('#2c3e50'), spaceAfter=30, alignment=TA_CENTER)

    agregar_logo_pdf(story, styles)
    story.append(Paragraph('Dashboard de Administrador', titulo_style))
    story.append(Paragraph(f'Administrador: {request.user.nombre_completo}', styles['Normal']))
    story.append(Paragraph('Popayán All Tour', styles['Normal']))
    story.append(Paragraph(f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    story.append(Spacer(1, 20))

    # Resumen general
    story.append(Paragraph('<b>Resumen General del Sistema</b>', styles['Heading2']))
    story.append(Spacer(1, 12))
    resumen_data = [
        ['Métrica', 'Valor'],
        ['Total Establecimientos', str(total_establecimientos)],
        ['Usuarios Registrados', str(usuarios_registrados)],
        ['Total Visitas', str(total_visitas)],
        ['Visitas este Mes', str(visitas_mes)],
    ]
    tabla_resumen = Table(resumen_data, colWidths=[3 * inch, 2 * inch])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(tabla_resumen)
    story.append(Spacer(1, 30))

    # Establecimientos por tipo
    story.append(Paragraph('<b>Establecimientos por Tipo</b>', styles['Heading2']))
    story.append(Spacer(1, 12))
    est_data = [['Tipo', 'Cantidad']] + [[k.title(), str(v)] for k, v in stats_establecimientos.items()]
    tabla_est = Table(est_data, colWidths=[3 * inch, 2 * inch])
    tabla_est.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(tabla_est)
    story.append(Spacer(1, 20))

    # Gráfica de establecimientos (Pie Chart)
    valores = list(stats_establecimientos.values())
    etiquetas = [k.title() for k in stats_establecimientos.keys()]
    colores_disponibles = [
        colors.HexColor('#3498db'), colors.HexColor('#e74c3c'),
        colors.HexColor('#f39c12'), colors.HexColor('#2ecc71'),
        colors.HexColor('#9b59b6'),
    ]
    drawing = Drawing(400, 200)
    pie = Pie()
    pie.x, pie.y, pie.width, pie.height = 150, 50, 100, 100
    pie.data = valores if valores else [1]
    pie.labels = etiquetas if etiquetas else ['Sin datos']
    for i in range(len(valores)):
        pie.slices[i].fillColor = colores_disponibles[i % len(colores_disponibles)]
    drawing.add(pie)
    story.append(drawing)
    story.append(Spacer(1, 30))

    # Usuarios por Rol
    story.append(Paragraph('<b>Usuarios por Rol</b>', styles['Heading2']))
    story.append(Spacer(1, 12))
    roles_tabla_data = [
        ['Rol', 'Cantidad'],
        ['Total Usuarios', str(usuarios_registrados)],
        ['Turistas', str(roles_dict.get('turista', 0))],
        ['Empresarios', str(roles_dict.get('empresario', 0))],
        ['Administradores', str(roles_dict.get('administrador', 0))],
    ]
    tabla_roles = Table(roles_tabla_data, colWidths=[3 * inch, 2 * inch])
    tabla_roles.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9b59b6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(tabla_roles)
    story.append(Spacer(1, 20))

    # Gráfica de barras: usuarios por rol
    datos_roles = [roles_dict.get('turista', 0), roles_dict.get('empresario', 0), roles_dict.get('administrador', 0)]
    drawing_roles = Drawing(400, 200)
    chart_roles = VerticalBarChart()
    chart_roles.x, chart_roles.y, chart_roles.height, chart_roles.width = 50, 50, 125, 300
    chart_roles.data = [datos_roles]
    chart_roles.categoryAxis.categoryNames = ['Turistas', 'Empresarios', 'Admins']
    chart_roles.valueAxis.valueMin = 0
    chart_roles.valueAxis.valueMax = (max(datos_roles) if datos_roles and max(datos_roles) > 0 else 1) * 1.2
    chart_roles.bars[0].fillColor = colors.HexColor('#3498db')
    drawing_roles.add(chart_roles)
    story.append(drawing_roles)
    story.append(Spacer(1, 30))

    # Distribución de visitas
    story.append(Paragraph('<b>Distribución de Visitas</b>', styles['Heading2']))
    story.append(Spacer(1, 12))
    visitas_data = [
        ['Tipo de Usuario', 'Visitas'],
        ['Usuarios Registrados', str(visitas_registradas)],
        ['Usuarios Invitados', str(visitas_invitados)],
    ]
    tabla_visitas = Table(visitas_data, colWidths=[3 * inch, 2 * inch])
    tabla_visitas.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9b59b6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(tabla_visitas)
    story.append(Spacer(1, 20))

    if visitas_registradas > 0 or visitas_invitados > 0:
        drawing2 = Drawing(400, 200)
        pie2 = Pie()
        pie2.x, pie2.y, pie2.width, pie2.height = 150, 50, 100, 100
        pie2.data = [visitas_registradas, visitas_invitados]
        pie2.labels = ['Registrados', 'Invitados']
        pie2.slices.strokeWidth = 0.5
        pie2.slices[0].fillColor = colors.HexColor('#2ecc71')
        pie2.slices[1].fillColor = colors.HexColor('#95a5a6')
        drawing2.add(pie2)
        story.append(drawing2)
    else:
        story.append(Paragraph('Sin datos de visitas disponibles.', styles['Normal']))
    story.append(Spacer(1, 30))

    # Top 5 establecimientos
    if top_establecimientos:
        story.append(Paragraph('<b>Top 5 Establecimientos Más Visitados</b>', styles['Heading2']))
        story.append(Spacer(1, 12))
        top_data = [['#', 'Nombre', 'Tipo', 'Visitas']] + [
            [str(i), item['nombre'], item['tipo'], str(item['visitas'])]
            for i, item in enumerate(top_establecimientos, start=1)
        ]
        tabla_top = Table(top_data, colWidths=[0.5 * inch, 2.5 * inch, 1.5 * inch, 1.5 * inch])
        tabla_top.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        story.append(tabla_top)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="dashboard_admin_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
def exportar_estadisticas_admin_excel(request):
    """Genera y descarga un Excel con estadísticas del administrador."""
    if not _es_admin(request.user):
        messages.error(request, 'No tienes permisos para esta acción.')
        return redirect('home')

    tipos = TipoEstablecimiento.objects.all()
    stats_establecimientos = {
        tipo.nombre.lower(): Establecimiento.objects.filter(tipo=tipo, activo=True).count()
        for tipo in tipos
    }
    total_establecimientos = sum(stats_establecimientos.values())
    usuarios_registrados = Usuario.objects.filter(is_active=True).count()
    total_visitas = visitas_mes = 0

    buffer = BytesIO()
    wb = Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen['A1'] = 'Dashboard de Administrador'
    ws_resumen['A1'].font = Font(size=18, bold=True, color='667eea')
    ws_resumen['A2'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

    row = 4
    for col in ['A', 'B']:
        cell = ws_resumen[f'{col}{row}']
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    ws_resumen[f'A{row}'] = 'Métrica'
    ws_resumen[f'B{row}'] = 'Valor'

    row += 1
    for metrica, valor in [
        ('Total Establecimientos', total_establecimientos),
        ('Usuarios Registrados', usuarios_registrados),
        ('Total Visitas', total_visitas),
        ('Visitas este Mes', visitas_mes),
    ]:
        ws_resumen[f'A{row}'] = metrica
        ws_resumen[f'B{row}'] = valor
        ws_resumen[f'B{row}'].alignment = Alignment(horizontal='center')
        row += 1

    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 15

    # ── Hoja 2: Establecimientos ──────────────────────────────────────────────
    ws_est = wb.create_sheet("Establecimientos")
    ws_est['A1'] = 'Tipo'
    ws_est['B1'] = 'Cantidad'
    for col in ['A', 'B']:
        ws_est[f'{col}1'].font = Font(bold=True, color='FFFFFF')
        ws_est[f'{col}1'].fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
        ws_est[f'{col}1'].alignment = Alignment(horizontal='center')

    for i, (tipo, cantidad) in enumerate(stats_establecimientos.items(), start=2):
        ws_est[f'A{i}'] = tipo.title()
        ws_est[f'B{i}'] = cantidad

    chart = PieChart()
    chart.title = "Establecimientos por Tipo"
    data = Reference(ws_est, min_col=2, min_row=1, max_row=len(stats_establecimientos) + 1)
    cats = Reference(ws_est, min_col=1, min_row=2, max_row=len(stats_establecimientos) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_est.add_chart(chart, "D2")
    ws_est.column_dimensions['A'].width = 20
    ws_est.column_dimensions['B'].width = 15

    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="dashboard_admin_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


# =============================================================================
# VISTAS — Exportación PDF/Excel (empresario)
# =============================================================================

@login_required
def exportar_estadisticas_empresario_pdf(request):
    """Genera y descarga un PDF con estadísticas del empresario y sus establecimientos."""
    if not hasattr(request.user, 'rol') or request.user.rol.rol.lower() != 'empresario':
        messages.error(request, 'No tienes permisos para esta acción.')
        return redirect('home')

    tipo_obj = request.user.tipo_establecimiento
    if not tipo_obj:
        messages.error(request, 'No tienes tipo de establecimiento asignado.')
        return redirect('home')

    establecimientos = Establecimiento.objects.filter(empresario=request.user, activo=True, tipo=tipo_obj)
    estadisticas = [
        {'nombre': e.nombre, 'visitas_totales': 0, 'visitas_mes': 0,
         'visitas_semana': 0, 'visitas_registradas': 0, 'visitas_anonimas': 0}
        for e in establecimientos
    ]
    total_visitas_general = 0

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    story = []
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24,
                                  textColor=colors.HexColor('#2c3e50'), spaceAfter=30, alignment=TA_CENTER)

    agregar_logo_pdf(story, styles)
    story.append(Paragraph(f'Estadísticas de {tipo_obj.nombre}', titulo_style))
    story.append(Paragraph(f'Empresario: {request.user.nombre_completo}', styles['Normal']))
    story.append(Paragraph(f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}', styles['Normal']))
    story.append(Spacer(1, 20))

    resumen_data = [
        ['Métrica', 'Valor'],
        ['Total de Establecimientos', str(establecimientos.count())],
        ['Visitas Totales Acumuladas', str(total_visitas_general)],
    ]
    tabla_resumen = Table(resumen_data, colWidths=[3 * inch, 2 * inch])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(tabla_resumen)
    story.append(Spacer(1, 30))

    tabla_data = [['Establecimiento', 'Total', 'Mes', 'Semana', 'Registrados', 'Anónimos']] + [
        [stat['nombre'], '0', '0', '0', '0', '0'] for stat in estadisticas
    ]
    tabla = Table(tabla_data, colWidths=[2 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 1 * inch, 1 * inch])
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(tabla)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="estadisticas_{tipo_obj.nombre}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    return response


@login_required
def exportar_estadisticas_empresario_excel(request):
    """Genera y descarga un Excel con estadísticas del empresario y sus establecimientos."""
    if not hasattr(request.user, 'rol') or request.user.rol.rol.lower() != 'empresario':
        messages.error(request, 'No tienes permisos para esta acción.')
        return redirect('home')

    tipo_obj = request.user.tipo_establecimiento
    if not tipo_obj:
        messages.error(request, 'No tienes tipo de establecimiento asignado.')
        return redirect('home')

    establecimientos = Establecimiento.objects.filter(empresario=request.user, activo=True, tipo=tipo_obj)

    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"

    ws['A1'] = f'Estadísticas de {tipo_obj.nombre}'
    ws['A1'].font = Font(size=18, bold=True, color='2c3e50')
    ws['A2'] = f'Empresario: {request.user.nombre_completo}'
    ws['A3'] = f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A5'] = 'Total de Establecimientos'
    ws['B5'] = establecimientos.count()
    ws['A6'] = 'Visitas Totales Acumuladas'
    ws['B6'] = 0

    headers = ['Establecimiento', 'Total', 'Mes', 'Semana', 'Registrados', 'Anónimos']
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=col_num, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2ecc71', end_color='2ecc71', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row_idx, est in enumerate(establecimientos, start=9):
        ws.cell(row=row_idx, column=1, value=est.nombre)
        for col in range(2, 7):
            ws.cell(row=row_idx, column=col, value=0)

    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 12

    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="estadisticas_{tipo_obj.nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    return response


# =============================================================================
# API — Endpoints JSON (dashboard y estadísticas)
# =============================================================================

@login_required
def api_estadisticas(request):
    """Retorna estadísticas generales en JSON para el dashboard."""
    if not _es_admin(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        tipos = TipoEstablecimiento.objects.all()
        stats = {
            'establecimientos': {
                tipo.nombre.lower(): Establecimiento.objects.filter(tipo=tipo, activo=True).count()
                for tipo in tipos
            },
            'usuarios': {
                'total': Usuario.objects.filter(is_active=True).count(),
                'turistas': Usuario.objects.filter(rol__rol='turista', is_active=True).count(),
                'empresarios': Usuario.objects.filter(rol__rol='empresario', is_active=True).count(),
            },
            'visitas': {'total': 0, 'registrados': 0, 'invitados': 0},
        }
        return JsonResponse(stats)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_establecimientos_visitas(request):
    """Retorna lista de establecimientos activos con sus visitas (actualmente en 0)."""
    if not _es_admin(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    try:
        data = [
            {'id': e.id, 'nombre': e.nombre, 'tipo': e.tipo.nombre, 'visitas': 0}
            for e in Establecimiento.objects.filter(activo=True).select_related('tipo')
        ]
        return JsonResponse({'success': True, 'establecimientos': data, 'total': len(data)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# =============================================================================
# API AJAX — Usuarios (dashboard administrador)
# =============================================================================

@login_required
@require_http_methods(['POST'])
def ajax_eliminar_usuario(request, user_id):
    """Elimina un usuario. No permite eliminar al propio admin ni al último admin activo."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        u = Usuario.objects.get(pk=user_id)
    except Usuario.DoesNotExist:
        return _json_error('Usuario no encontrado.', 404)

    if u.pk == request.user.pk:
        return _json_error('No puedes eliminarte a ti mismo.')

    if hasattr(u, 'rol') and u.rol and u.rol.rol.lower() == 'administrador':
        admins_restantes = Usuario.objects.filter(
            rol__rol__iexact='administrador', is_active=True
        ).exclude(pk=u.pk).count()
        if admins_restantes == 0:
            return _json_error('No puedes eliminar al único administrador activo.')

    nombre = u.nombre_completo
    u.delete()
    return _json_ok(msg=f'Usuario "{nombre}" eliminado correctamente.')


@login_required
@require_http_methods(['POST'])
def ajax_suspender_usuario(request, user_id):
    """Desactiva un usuario vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        u = Usuario.objects.get(pk=user_id)
    except Usuario.DoesNotExist:
        return _json_error('Usuario no encontrado.', 404)

    if u.pk == request.user.pk:
        return _json_error('No puedes suspenderte a ti mismo.')

    u.is_active = False
    u.save(update_fields=['is_active'])
    return _json_ok(msg=f'Usuario "{u.nombre_completo}" suspendido.')


@login_required
@require_http_methods(['POST'])
def ajax_activar_usuario(request, user_id):
    """Activa un usuario vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        u = Usuario.objects.get(pk=user_id)
    except Usuario.DoesNotExist:
        return _json_error('Usuario no encontrado.', 404)

    u.is_active = True
    u.save(update_fields=['is_active'])
    return _json_ok(msg=f'Usuario "{u.nombre_completo}" activado.')


@login_required
@require_http_methods(['POST'])
def ajax_editar_usuario(request, user_id):
    """Edita rol, tipo_establecimiento e is_active de un usuario vía AJAX (JSON o form-data)."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        u = Usuario.objects.select_related('rol', 'tipo_establecimiento').get(pk=user_id)
    except Usuario.DoesNotExist:
        return _json_error('Usuario no encontrado.', 404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    rol_id = data.get('rol')
    tipo_id = data.get('tipo_establecimiento') or None
    activo = data.get('is_active')

    if rol_id:
        try:
            u.rol = Roles.objects.get(pk=rol_id)
        except Roles.DoesNotExist:
            return _json_error('Rol no encontrado.', 404)

    if tipo_id:
        try:
            u.tipo_establecimiento = TipoEstablecimiento.objects.get(pk=tipo_id)
        except TipoEstablecimiento.DoesNotExist:
            return _json_error('Tipo no encontrado.', 404)
    else:
        u.tipo_establecimiento = None

    if activo is not None:
        u.is_active = str(activo).lower() in ('true', '1', 'yes')

    u.save(update_fields=['rol', 'tipo_establecimiento', 'is_active'])
    return _json_ok(
        data={
            'usuario': {
                'id': u.pk,
                'nombre': u.nombre_completo,
                'email': u.email,
                'rol': u.rol.rol if u.rol else '—',
                'tipo': u.tipo_establecimiento.nombre if u.tipo_establecimiento else '—',
                'is_active': u.is_active,
            }
        },
        msg='Usuario actualizado correctamente.',
    )


# =============================================================================
# API AJAX — Establecimientos (dashboard administrador)
# =============================================================================

@login_required
@require_http_methods(['POST'])
def ajax_eliminar_establecimiento(request, estab_id):
    """Soft delete de un establecimiento vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        e = Establecimiento.objects.select_related('tipo').get(pk=estab_id)
    except Establecimiento.DoesNotExist:
        return _json_error('Establecimiento no encontrado.', 404)

    nombre = e.nombre
    e.activo = False
    e.save(update_fields=['activo'])
    return _json_ok(msg=f'Sitio "{nombre}" desactivado correctamente.')


@login_required
@require_http_methods(['POST'])
def ajax_activar_establecimiento(request, estab_id):
    """Reactiva un establecimiento vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        e = Establecimiento.objects.get(pk=estab_id)
    except Establecimiento.DoesNotExist:
        return _json_error('Establecimiento no encontrado.', 404)

    e.activo = True
    e.save(update_fields=['activo'])
    return _json_ok(msg=f'Sitio "{e.nombre}" activado correctamente.')


# =============================================================================
# API AJAX — Stats en tiempo real
# =============================================================================

@login_required
def ajax_dashboard_stats(request):
    """Retorna métricas actualizadas del dashboard en JSON (para el botón Actualizar)."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    tipos = TipoEstablecimiento.objects.annotate(
        total=Count('establecimientos', filter=Q(establecimientos__activo=True))
    )
    stats_est = {t.nombre.lower(): t.total for t in tipos}

    roles_data = {
        (i['rol__rol'] or 'sin rol').lower(): i['total']
        for i in (
            Usuario.objects.filter(is_active=True)
            .values('rol__rol')
            .annotate(total=Count('id'))
        )
    }

    return JsonResponse({
        'success': True,
        'stats': {
            'total_establecimientos': sum(stats_est.values()),
            'usuarios_registrados': Usuario.objects.filter(is_active=True).count(),
            'usuarios_suspendidos': Usuario.objects.filter(is_active=False).count(),
            'noticias_publicadas': Noticia.objects.filter(publicada=True).count(),
            'noticias_borrador': Noticia.objects.filter(publicada=False).count(),
            'noticias_destacadas': Noticia.objects.filter(destacada=True).count(),
            'roles': roles_data,
            'establecimientos_por_tipo': stats_est,
        },
    })

#cambios noticias/dashboard
# =============================================================================
# API AJAX — Noticias (dashboard administrador)
# =============================================================================

@login_required
@require_http_methods(['POST'])
def ajax_suspender_noticia(request, slug):
    """Suspende una noticia (publicada=False) vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        n = Noticia.objects.get(slug=slug)
    except Noticia.DoesNotExist:
        return _json_error('Noticia no encontrada.', 404)

    n.publicada = False
    n.save(update_fields=['publicada'])
    return _json_ok(msg=f'Noticia "{n.titulo}" suspendida.')


@login_required
@require_http_methods(['POST'])
def ajax_reactivar_noticia(request, slug):
    """Reactiva una noticia (publicada=True) vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        n = Noticia.objects.get(slug=slug)
    except Noticia.DoesNotExist:
        return _json_error('Noticia no encontrada.', 404)

    n.publicada = True
    n.save(update_fields=['publicada'])
    return _json_ok(msg=f'Noticia "{n.titulo}" reactivada.')


@login_required
@require_http_methods(['POST'])
def ajax_eliminar_noticia_definitivo(request, slug):
    """Elimina definitivamente una noticia vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        n = Noticia.objects.get(slug=slug)
    except Noticia.DoesNotExist:
        return _json_error('Noticia no encontrada.', 404)

    titulo = n.titulo
    n.delete()
    return _json_ok(msg=f'Noticia "{titulo}" eliminada definitivamente.')

#catorias noticias/dashboard
# Reemplazar la función crear_categoria existente:

@login_required
@require_http_methods(['GET', 'POST'])
def crear_categoria(request):
    """Crea una nueva categoría de noticias vía AJAX (JSON) o form POST."""
    if not (
        request.user.is_superuser or
        (hasattr(request.user, 'rol') and request.user.rol.rol.lower() == 'administrador')
    ):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
            return _json_error('No tienes permisos para crear categorías', 403)
        messages.error(request, 'No tienes permisos para crear categorías')
        return redirect('home')

    if request.method == 'POST':
        form = CategoriaNoticiaForm(request.POST)
        if form.is_valid():
            categoria = form.save()
            return _json_ok(
                data={
                    'categoria': {
                        'id': categoria.id,
                        'nombre': categoria.nombre,
                        'descripcion': categoria.descripcion or '',
                        'color': categoria.color,
                        'activo': categoria.activo,
                        'noticias_count': 0,
                    }
                },
                msg=f'Categoría "{categoria.nombre}" creada exitosamente.',
            )
        else:
            errores = {field: errors[0] for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errores}, status=400)

    return redirect('dashboard_administrador')


# Agregar estas 3 funciones al final de views.py:

@login_required
@require_http_methods(['POST'])
def ajax_editar_categoria(request, cat_id):
    """Edita nombre, descripción y color de una categoría vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        c = CategoriaNoticia.objects.get(pk=cat_id)
    except CategoriaNoticia.DoesNotExist:
        return _json_error('Categoría no encontrada.', 404)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        data = request.POST

    nombre = data.get('nombre', '').strip()
    if not nombre:
        return _json_error('El nombre no puede estar vacío.')

    c.nombre = nombre
    c.descripcion = data.get('descripcion', '').strip()
    color = data.get('color', '').strip()
    if color:
        c.color = color
    c.save(update_fields=['nombre', 'descripcion', 'color'])

    return _json_ok(
        data={
            'categoria': {
                'id': c.id,
                'nombre': c.nombre,
                'descripcion': c.descripcion or '',
                'color': c.color,
                'activo': c.activo,
            }
        },
        msg=f'Categoría "{c.nombre}" actualizada.',
    )


@login_required
@require_http_methods(['POST'])
def ajax_toggle_categoria(request, cat_id):
    """Activa o suspende una categoría vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        c = CategoriaNoticia.objects.get(pk=cat_id)
    except CategoriaNoticia.DoesNotExist:
        return _json_error('Categoría no encontrada.', 404)

    c.activo = not c.activo
    c.save(update_fields=['activo'])
    estado = 'activada' if c.activo else 'suspendida'
    return _json_ok(
        data={'activo': c.activo},
        msg=f'Categoría "{c.nombre}" {estado}.',
    )


@login_required
@require_http_methods(['POST'])
def ajax_eliminar_categoria(request, cat_id):
    """Elimina definitivamente una categoría vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)

    try:
        c = CategoriaNoticia.objects.get(pk=cat_id)
    except CategoriaNoticia.DoesNotExist:
        return _json_error('Categoría no encontrada.', 404)

    if c.noticias.exists():
        return _json_error(
            f'No se puede eliminar: tiene {c.noticias.count()} noticia(s) asociada(s). Reasígnalas primero.'
        )

    nombre = c.nombre
    c.delete()
    return _json_ok(msg=f'Categoría "{nombre}" eliminada definitivamente.')

#news funciones dashboard resenas
# =============================================================================
# API AJAX — Reseñas (dashboard administrador)
# =============================================================================

@login_required
def ajax_lista_resenas(request):
    """Retorna todas las reseñas en JSON para el dashboard."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)
    
    resenas = (
        Resena.objects
        .select_related('usuario', 'establecimiento', 'establecimiento__tipo')
        .order_by('-fecha_creacion')
    )
    data = [
        {
            'id': r.id,
            'usuario_nombre': r.usuario.nombre_completo,
            'usuario_email': r.usuario.email,
            'establecimiento': r.establecimiento.nombre,
            'tipo': r.establecimiento.tipo.nombre if r.establecimiento.tipo else '—',
            'calificacion': r.calificacion,
            'comentario': r.comentario or '',
            'fecha': r.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
        }
        for r in resenas
    ]
    return JsonResponse({'success': True, 'resenas': data, 'total': len(data)})


@login_required
@require_http_methods(['POST'])
def ajax_eliminar_resena(request, resena_id):
    """Elimina una reseña definitivamente vía AJAX."""
    if not _es_admin(request.user):
        return _json_error('Sin permisos.', 403)
    
    try:
        r = Resena.objects.select_related('usuario', 'establecimiento').get(pk=resena_id)
    except Resena.DoesNotExist:
        return _json_error('Reseña no encontrada.', 404)
    
    nombre_usuario = r.usuario.nombre_completo
    nombre_estab = r.establecimiento.nombre
    r.delete()
    return _json_ok(msg=f'Reseña de "{nombre_usuario}" en "{nombre_estab}" eliminada.')